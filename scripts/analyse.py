"""
analyse.py — Compute confirmed metrics and generate reports.

All calculations use raw granular data (individual time entries, deal-level
financials). No Productive built-in aggregates are used.

Metric definitions were confirmed with the user on 2026-03-19.

Usage:
    python scripts/analyse.py --report financial
    python scripts/analyse.py --report people
    python scripts/analyse.py --report project
    python scripts/analyse.py --report client
    python scripts/analyse.py --report hygiene
    python scripts/analyse.py --report all
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

sys.path.insert(0, str(Path(__file__).parent))
from transform import transform_endpoint, find_latest_snapshot, load_snapshot
from load_budget import load_all_budgets, match_budget_name

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# ============================================================
# Color constants (from skills/visualization.md)
# ============================================================

COLORS = {
    "billable": "#2ECC71",
    "non_billable": "#95A5A6",
    "revenue": "#3498DB",
    "cost": "#E67E22",
    "overbudget": "#E74C3C",
    "warning": "#F39C12",
    "on_track": "#27AE60",
    "profit_pos": "#27AE60",
    "profit_neg": "#E74C3C",
}

CHART_LAYOUT = dict(
    template="plotly_white",
    font=dict(size=12),
    title_font_size=16,
    hoverlabel=dict(font_size=12),
    margin=dict(l=60, r=40, t=60, b=60),
)


def _save_chart(fig: go.Figure, name: str):
    """Save plotly figure as interactive HTML."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / f"{name}.html"
    fig.write_html(str(path), include_plotlyjs="cdn")
    print(f"  Chart saved to {path}")


# ============================================================
# Metric annotations — always clarify what numbers mean
# ============================================================

ANNOTATIONS = {
    "revenue": (
        "Revenue (Recognized): from Productive financial_item_reports API "
        "(total_recognized_revenue grouped by month). Matches Productive's "
        "own reports exactly."
    ),
    "staff_cost": (
        "Staff Cost: salary-based cost from time entries (work_cost). "
        "Excludes overhead. Owners (Johan VDC, Johan VT, Michel) "
        "overridden to EUR 14,166.67/month flat (EUR 42,500/3)."
    ),
    "gross_margin": (
        "Gross Margin (excl. overhead) = Revenue - Staff Cost. "
        "Does NOT include facility or internal overhead costs."
    ),
    "utilisation": (
        "Billable Utilisation = Billable Hours / Total Tracked Hours. "
        "Denominator is all logged time (billable + non-billable/internal). "
        "Contractors may not track full days."
    ),
    "overbudget": (
        "Overbudget flags: Green <70%, Amber 70-100%, Red >100%. "
        "Dual check on hours AND EUR — either triggers the flag."
    ),
    "acph": (
        "ACPH (Avg Cost Per Hour) = Staff Cost / Total Hours by service type. "
        "ARPH (Avg Revenue Per Hour) = Revenue / Billable Hours by service type."
    ),
}


def _load_data(endpoints: list[str]) -> dict[str, pd.DataFrame]:
    """Load only the endpoints needed for a report."""
    data = {}
    for ep in endpoints:
        df = transform_endpoint(ep)
        if df is not None:
            data[ep] = df
    return data


def _save_report(df: pd.DataFrame, name: str):
    """Save report DataFrame to CSV."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / f"{name}.csv"
    df.to_csv(path, index=False)
    print(f"  Saved to {path}")


def load_financial_item_reports() -> pd.DataFrame | None:
    """Load financial_item_reports snapshot and return a DataFrame with monthly revenue.

    Each record has date_period like "2026/M01" and financial totals in cents.
    Returns DataFrame with columns: month_str, revenue, invoiced, cost, profit.
    """
    filepath = find_latest_snapshot("financial_item_reports")
    if not filepath:
        return None

    snapshot = load_snapshot(filepath)
    records = []
    for item in snapshot.get("data", []):
        attrs = item.get("attributes", {})
        date_period = attrs.get("date_period", "")
        if not date_period or "/M" not in date_period:
            continue
        # "2026/M01" → "2026-01"
        year, month_part = date_period.split("/M")
        month_str = f"{year}-{month_part}"
        records.append({
            "month_str": month_str,
            "revenue": float(attrs.get("total_recognized_revenue", 0)) / 100.0,
            "invoiced": float(attrs.get("total_invoiced", 0)) / 100.0,
            "cost": float(attrs.get("total_cost", 0)) / 100.0,
            "profit": float(attrs.get("total_recognized_profit", 0)) / 100.0,
        })

    if not records:
        return None

    df = pd.DataFrame(records).sort_values("month_str").reset_index(drop=True)
    print(f"  Loaded {len(df)} months of recognized revenue from financial_item_reports")
    return df


def load_financial_item_reports_by_budget() -> pd.DataFrame | None:
    """Load per-budget financial_item_reports and return revenue per budget per month.

    Uses the snapshot fetched with group=date:month,budget.
    Budget ID is extracted from the record ID via regex.
    Returns DataFrame with columns: budget_id, month_str, revenue, invoiced.
    """
    import re

    filepath = find_latest_snapshot("financial_item_reports_by_budget")
    if not filepath:
        return None

    snapshot = load_snapshot(filepath)
    records = []
    for item in snapshot.get("data", []):
        attrs = item.get("attributes", {})
        date_period = attrs.get("date_period", "")
        if not date_period or "/M" not in date_period:
            continue

        # Extract budget_id from record ID, e.g. "...-budget-12345-..."
        record_id = item.get("id", "")
        match = re.search(r"budget-(\d+)", record_id)
        if not match:
            continue
        budget_id = match.group(1)

        # "2026/M01" → "2026-01"
        year, month_part = date_period.split("/M")
        month_str = f"{year}-{month_part}"

        records.append({
            "budget_id": budget_id,
            "month_str": month_str,
            "revenue": float(attrs.get("total_recognized_revenue", 0)) / 100.0,
            "invoiced": float(attrs.get("total_invoiced", 0)) / 100.0,
        })

    if not records:
        return None

    df = pd.DataFrame(records)
    print(f"  Loaded {len(df)} per-budget revenue records from financial_item_reports_by_budget")
    return df


def load_financial_item_reports_by_service_type() -> pd.DataFrame | None:
    """Load per-service-type financial_item_reports and return revenue per service_type per month.

    Uses the snapshot fetched with group=date:month,service_type and include=service_type.
    Service type names are resolved from the included entities.
    Returns DataFrame with columns: service_type, month_str, revenue, cost.
    """
    filepath = find_latest_snapshot("financial_item_reports_by_service_type")
    if not filepath:
        return None

    snapshot = load_snapshot(filepath)

    # Build service_type ID → name lookup from included entities
    st_lookup = {}
    for inc in snapshot.get("included", []):
        if inc.get("type") == "service_types":
            st_lookup[inc["id"]] = inc.get("attributes", {}).get("name", f"id:{inc['id']}")

    records = []
    for item in snapshot.get("data", []):
        attrs = item.get("attributes", {})
        date_period = attrs.get("date_period", "")
        if not date_period or "/M" not in date_period:
            continue

        # Resolve service_type name from relationship
        st_rel = item.get("relationships", {}).get("service_type", {}).get("data")
        if st_rel:
            st_name = st_lookup.get(st_rel["id"], f"id:{st_rel['id']}")
        else:
            st_name = "(no type)"

        # "2026/M01" → "2026-01"
        year, month_part = date_period.split("/M")
        month_str = f"{year}-{month_part}"

        records.append({
            "service_type": st_name,
            "month_str": month_str,
            "revenue": float(attrs.get("total_recognized_revenue", 0)) / 100.0,
            "cost": float(attrs.get("total_cost", 0)) / 100.0,
        })

    if not records:
        return None

    df = pd.DataFrame(records)
    # Aggregate in case multiple records map to the same service_type + month
    df = df.groupby(["service_type", "month_str"]).agg(
        revenue=("revenue", "sum"),
        cost=("cost", "sum"),
    ).reset_index()
    print(f"  Loaded {len(df)} per-service-type revenue records from financial_item_reports_by_service_type")
    return df


def _build_budget_to_deal_map() -> dict[str, dict]:
    """Build budget_id → {deal_name, deal_id, company_name, service_type} from deals snapshot.

    In Productive, budgets ARE deals (budget=True). The deal ID is the budget ID.
    Service type comes from the services snapshot (service → deal relationship).
    """
    # Load deals to get budget_id → deal_name, company_name
    deals_filepath = find_latest_snapshot("deals")
    if not deals_filepath:
        return {}

    deals_snapshot = load_snapshot(deals_filepath)
    lookup = {}
    included = {(i["type"], i["id"]): i for i in deals_snapshot.get("included", [])}

    for item in deals_snapshot.get("data", []):
        deal_id = item["id"]
        attrs = item.get("attributes", {})
        name = attrs.get("name", "")

        # Get company name and ID from relationship
        company_name = None
        company_id = None
        co_ref = item.get("relationships", {}).get("company", {}).get("data")
        if co_ref:
            company_id = co_ref["id"]
            co = included.get((co_ref["type"], co_ref["id"]))
            if co:
                company_name = co.get("attributes", {}).get("name", "")

        # Get project ID from relationship
        project_id = None
        proj_ref = item.get("relationships", {}).get("project", {}).get("data")
        if proj_ref:
            project_id = proj_ref["id"]

        lookup[deal_id] = {
            "deal_name": name,
            "deal_id": deal_id,
            "company_name": company_name,
            "company_id": company_id,
            "project_id": project_id,
        }

    # Enrich with primary service_type from services snapshot
    svc_filepath = find_latest_snapshot("services")
    if svc_filepath:
        svc_snapshot = load_snapshot(svc_filepath)
        svc_included = {(i["type"], i["id"]): i for i in svc_snapshot.get("included", [])}

        # Collect service_type per deal (use the most common one)
        deal_service_types: dict[str, list[str]] = {}
        for svc in svc_snapshot.get("data", []):
            deal_ref = svc.get("relationships", {}).get("deal", {}).get("data")
            if not deal_ref:
                continue
            did = deal_ref["id"]
            st_ref = svc.get("relationships", {}).get("service_type", {}).get("data")
            if st_ref:
                st = svc_included.get((st_ref["type"], st_ref["id"]))
                if st:
                    st_name = st.get("attributes", {}).get("name", "")
                    if st_name:
                        deal_service_types.setdefault(did, []).append(st_name)

        for did, types in deal_service_types.items():
            if did in lookup:
                # Most common service type for this deal
                from collections import Counter
                lookup[did]["service_type"] = Counter(types).most_common(1)[0][0]

    return lookup


# ============================================================
# 1. FINANCIAL REPORT
# ============================================================

def report_financial(te: pd.DataFrame, deals: pd.DataFrame) -> pd.DataFrame:
    """
    Revenue, Staff Cost, Gross Margin by month. YTD, YoY, EOY forecast.

    Revenue = deal.revenue (recognized, per Productive settings)
    Staff Cost = work_cost (salary-based, no overhead, owners = EUR 15K/month)
    Gross Margin = Revenue - Staff Cost
    """
    print("\n" + "=" * 70)
    print("FINANCIAL REPORT")
    print("=" * 70)
    print(f"\n  {ANNOTATIONS['revenue']}")
    print(f"  {ANNOTATIONS['staff_cost']}")
    print(f"  {ANNOTATIONS['gross_margin']}\n")

    # --- Monthly from time entries (staff cost + hours) ---
    monthly_te = te.groupby("month").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
    ).reset_index()
    monthly_te["month_str"] = monthly_te["month"].astype(str)

    # --- Revenue from financial_item_reports API ---
    fir = load_financial_item_reports()
    if fir is not None:
        monthly_te = monthly_te.merge(
            fir[["month_str", "revenue"]], on="month_str", how="left"
        )
        monthly_te["revenue"] = monthly_te["revenue"].fillna(0)
    else:
        print("  WARNING: No financial_item_reports snapshot found, revenue will be 0")
        monthly_te["revenue"] = 0

    client_budgets = deals[
        (deals["budget"] == True) & (deals["deal_type_id"] == 2)
    ].copy()
    open_deals = client_budgets[client_budgets["delivered_on"].isna()]
    delivered_deals = client_budgets[client_budgets["delivered_on"].notna()]

    total_revenue = monthly_te["revenue"].sum()
    total_staff_cost = monthly_te["staff_cost"].sum()
    total_gross_margin = total_revenue - total_staff_cost
    margin_pct = (total_gross_margin / total_revenue * 100) if total_revenue > 0 else 0

    print("TOTALS (all time):")
    print(f"  Revenue (Recognized):    EUR {total_revenue:>12,.2f}")
    print(f"  Staff Cost:              EUR {total_staff_cost:>12,.2f}")
    print(f"  Gross Margin:            EUR {total_gross_margin:>12,.2f} ({margin_pct:.1f}%)")
    print(f"  Client budgets:          {len(client_budgets)} ({len(open_deals)} open, {len(delivered_deals)} delivered)")

    # --- YTD ---
    current_year = pd.Timestamp.now().year
    ytd = monthly_te[monthly_te["month_str"].str.startswith(str(current_year))]
    if not ytd.empty:
        ytd_rev = ytd["revenue"].sum()
        ytd_cost = ytd["staff_cost"].sum()
        ytd_hours = ytd["total_hours"].sum()
        ytd_billable = ytd["billable_hours"].sum()
        ytd_util = (ytd_billable / ytd_hours * 100) if ytd_hours > 0 else 0
        print(f"\nYTD {current_year}:")
        print(f"  Revenue:        EUR {ytd_rev:>12,.2f}")
        print(f"  Staff Cost:     EUR {ytd_cost:>12,.2f}")
        print(f"  Gross Margin:   EUR {ytd_rev - ytd_cost:>12,.2f}")
        print(f"  Total Hours:    {ytd_hours:>12,.1f}h")
        print(f"  Billable Hours: {ytd_billable:>12,.1f}h ({ytd_util:.1f}%)")

    # --- YoY comparison ---
    prev_year = current_year - 1
    current_month = pd.Timestamp.now().month
    prev_ytd = monthly_te[
        monthly_te["month_str"].str.startswith(str(prev_year))
        & (monthly_te["month"].apply(lambda m: m.month) <= current_month)
    ]
    if not prev_ytd.empty:
        print(f"\nYoY (Jan-{pd.Timestamp.now().strftime('%b')}):")
        print(f"  {current_year} Staff Cost: EUR {ytd_cost:>10,.0f} | Hours: {ytd_hours:>8,.1f}h")
        prev_cost = prev_ytd["staff_cost"].sum()
        prev_hours = prev_ytd["total_hours"].sum()
        print(f"  {prev_year} Staff Cost: EUR {prev_cost:>10,.0f} | Hours: {prev_hours:>8,.1f}h")

    # --- EOY Forecast ---
    # Trailing 3 complete months average, with H2 2025 seasonality
    complete_months = monthly_te[
        monthly_te["month_str"] < pd.Timestamp.now().strftime("%Y-%m")
    ]
    if len(complete_months) >= 3:
        trailing_3 = complete_months.tail(3)
        avg_monthly_cost = trailing_3["staff_cost"].mean()
        avg_monthly_hours = trailing_3["total_hours"].mean()

        # Seasonality from H2 2025
        h2_2025 = monthly_te[
            (monthly_te["month_str"] >= "2025-07")
            & (monthly_te["month_str"] <= "2025-12")
        ]
        if not h2_2025.empty:
            h2_avg = h2_2025["staff_cost"].mean()
            # Build monthly seasonality factors
            seasonality = {}
            for _, row in h2_2025.iterrows():
                m = row["month"].month
                seasonality[m] = row["staff_cost"] / h2_avg if h2_avg > 0 else 1.0

        remaining_months = 12 - current_month + 1  # include current partial
        forecast_cost = ytd_cost
        for m in range(current_month + 1, 13):
            factor = seasonality.get(m, 1.0) if seasonality else 1.0
            forecast_cost += avg_monthly_cost * factor

        print(f"\nEOY Forecast {current_year}:")
        print(f"  Trailing 3-month avg cost: EUR {avg_monthly_cost:>10,.0f}/month")
        print(f"  Forecasted Staff Cost:     EUR {forecast_cost:>10,.0f}")
        print(f"  (Based on trailing 3-month avg with H2 2025 seasonality)")

    # --- Monthly breakdown ---
    monthly_te["util_pct"] = (
        monthly_te["billable_hours"] / monthly_te["total_hours"] * 100
    ).round(1)
    print(f"\nMonthly Breakdown:")
    print(f"  {'Month':<10} {'Revenue':>12} {'Staff Cost':>14} {'Margin':>12} {'Hours':>8} {'Util%':>7}")
    print(f"  {'-'*10} {'-'*12} {'-'*14} {'-'*12} {'-'*8} {'-'*7}")
    for _, row in monthly_te.iterrows():
        gm = row["revenue"] - row["staff_cost"]
        print(
            f"  {row['month_str']:<10} EUR {row['revenue']:>8,.0f} "
            f"EUR {row['staff_cost']:>10,.0f} EUR {gm:>8,.0f} "
            f"{row['total_hours']:>8,.1f} {row['util_pct']:>6.1f}%"
        )

    _save_report(monthly_te, "financial_monthly")

    # --- Charts ---
    charts = _charts_financial(monthly_te, te, deals, client_budgets)

    return monthly_te, charts


def _charts_financial(
    monthly_te: pd.DataFrame,
    te: pd.DataFrame,
    deals: pd.DataFrame,
    client_budgets: pd.DataFrame,
) -> dict[str, go.Figure]:
    """Build financial report charts. Returns {name: figure}."""
    charts = {}
    current_year = pd.Timestamp.now().year
    prev_year = current_year - 1
    current_month_str = pd.Timestamp.now().strftime("%Y-%m")

    # monthly_te already has revenue from report_financial()
    monthly = monthly_te.copy()
    if "revenue" not in monthly.columns:
        monthly["revenue"] = 0

    monthly["gross_margin"] = monthly["revenue"] - monthly["staff_cost"]
    monthly["margin_pct"] = (
        monthly["gross_margin"] / monthly["revenue"].replace(0, float("nan")) * 100
    ).round(1)

    # 1. Monthly Revenue & Staff Cost — grouped bar with margin line + data labels
    many = len(monthly) > 6
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    if many:
        fig.add_trace(
            go.Bar(
                x=monthly["month_str"], y=monthly["revenue"],
                name="Revenue", marker_color=COLORS["revenue"],
                hovertemplate="Revenue: €%{y:,.0f}<extra></extra>",
            ), secondary_y=False,
        )
        fig.add_trace(
            go.Bar(
                x=monthly["month_str"], y=monthly["staff_cost"],
                name="Staff Cost", marker_color=COLORS["cost"],
                hovertemplate="Staff Cost: €%{y:,.0f}<extra></extra>",
            ), secondary_y=False,
        )
        fig.add_trace(
            go.Scatter(
                x=monthly["month_str"], y=monthly["margin_pct"],
                name="Gross Margin %", mode="lines+markers",
                line=dict(color=COLORS["profit_pos"], width=3),
                marker=dict(size=6, color=COLORS["profit_pos"],
                            line=dict(width=2, color="white")),
                hovertemplate="Margin: %{y:.1f}%<extra></extra>",
            ), secondary_y=True,
        )
    else:
        fig.add_trace(
            go.Bar(
                x=monthly["month_str"], y=monthly["revenue"],
                name="Revenue", marker_color=COLORS["revenue"],
                text=[f"€{v:,.0f}" for v in monthly["revenue"]],
                textposition="outside",
                textfont=dict(size=10, color="#2C3E50"),
                cliponaxis=False,
                hovertemplate="Revenue: €%{y:,.0f}<extra></extra>",
            ), secondary_y=False,
        )
        fig.add_trace(
            go.Bar(
                x=monthly["month_str"], y=monthly["staff_cost"],
                name="Staff Cost", marker_color=COLORS["cost"],
                text=[f"€{v:,.0f}" for v in monthly["staff_cost"]],
                textposition="outside",
                textfont=dict(size=10, color="#2C3E50"),
                cliponaxis=False,
                hovertemplate="Staff Cost: €%{y:,.0f}<extra></extra>",
            ), secondary_y=False,
        )
        fig.add_trace(
            go.Scatter(
                x=monthly["month_str"], y=monthly["margin_pct"],
                name="Gross Margin %", mode="lines+markers+text",
                line=dict(color=COLORS["profit_pos"], width=3),
                marker=dict(size=10, color=COLORS["profit_pos"],
                            line=dict(width=2, color="white")),
                text=[f"{v:.1f}%" for v in monthly["margin_pct"]],
                textposition="top center",
                textfont=dict(size=11, color=COLORS["profit_pos"]),
                hovertemplate="Margin: %{y:.1f}%<extra></extra>",
            ), secondary_y=True,
        )
        for _, row in monthly.iterrows():
            gm_color = COLORS["profit_pos"] if row["gross_margin"] >= 0 else COLORS["profit_neg"]
            fig.add_annotation(
                x=row["month_str"], y=0, yref="y",
                text=f"<b>GM €{row['gross_margin']:,.0f}</b>",
                showarrow=False, yshift=-28,
                font=dict(size=10, color=gm_color),
            )

    y_max = monthly["revenue"].max() * (1.08 if many else 1.35)
    layout_overrides = dict(CHART_LAYOUT)
    layout_overrides["margin"] = dict(l=70, r=50, t=80, b=50 if many else 80)
    fig.update_layout(
        **layout_overrides,
        title="Monthly Revenue & Staff Cost",
        barmode="group",
        bargap=0.2,
        bargroupgap=0.05,
        yaxis_title="EUR",
        yaxis2_title="Gross Margin %",
        yaxis=dict(
            range=[0, y_max],
            tickprefix="€",
            tickformat=",",
        ),
        yaxis2=dict(
            ticksuffix="%",
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="center", x=0.5),
        height=550,
    )
    charts["financial_monthly_revenue_cost"] = fig

    # 2. YoY Comparison — overlaid line chart (2025 vs 2026)
    for year in [prev_year, current_year]:
        year_data = monthly[monthly["month_str"].str.startswith(str(year))].copy()
        if not year_data.empty:
            year_data["month_num"] = year_data["month"].apply(lambda m: m.month)

    curr = monthly[monthly["month_str"].str.startswith(str(current_year))].copy()
    prev = monthly[monthly["month_str"].str.startswith(str(prev_year))].copy()
    if not curr.empty and not prev.empty:
        curr["month_num"] = curr["month"].apply(lambda m: m.month)
        prev["month_num"] = prev["month"].apply(lambda m: m.month)
        month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        fig_yoy = go.Figure()
        fig_yoy.add_trace(go.Scatter(
            x=prev["month_num"], y=prev["staff_cost"],
            name=f"{prev_year} Staff Cost", mode="lines+markers",
            line=dict(color=COLORS["cost"], dash="dash"),
            hovertemplate=f"{prev_year}: EUR %{{y:,.0f}}<extra></extra>",
        ))
        fig_yoy.add_trace(go.Scatter(
            x=curr["month_num"], y=curr["staff_cost"],
            name=f"{current_year} Staff Cost", mode="lines+markers",
            line=dict(color=COLORS["cost"]),
            hovertemplate=f"{current_year}: EUR %{{y:,.0f}}<extra></extra>",
        ))
        fig_yoy.add_trace(go.Scatter(
            x=prev["month_num"], y=prev["revenue"],
            name=f"{prev_year} Revenue", mode="lines+markers",
            line=dict(color=COLORS["revenue"], dash="dash"),
            hovertemplate=f"{prev_year}: EUR %{{y:,.0f}}<extra></extra>",
        ))
        fig_yoy.add_trace(go.Scatter(
            x=curr["month_num"], y=curr["revenue"],
            name=f"{current_year} Revenue", mode="lines+markers",
            line=dict(color=COLORS["revenue"]),
            hovertemplate=f"{current_year}: EUR %{{y:,.0f}}<extra></extra>",
        ))
        fig_yoy.update_layout(
            **CHART_LAYOUT,
            title=f"Year-over-Year: {prev_year} vs {current_year}",
            xaxis=dict(
                tickvals=list(range(1, 13)),
                ticktext=month_labels,
                title="Month",
            ),
            yaxis_title="EUR",
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        charts["financial_yoy"] = fig_yoy

    # 3. EOY Forecast — bar chart with solid actuals + hatched forecast
    complete = monthly[monthly["month_str"] < current_month_str].copy()
    if len(complete) >= 3:
        trailing_3 = complete.tail(3)
        avg_cost = trailing_3["staff_cost"].mean()
        avg_rev = trailing_3["revenue"].mean()

        # Seasonality from H2 2025
        h2 = monthly[
            (monthly["month_str"] >= "2025-07") & (monthly["month_str"] <= "2025-12")
        ]
        seasonality = {}
        if not h2.empty:
            h2_avg = h2["staff_cost"].mean()
            for _, row in h2.iterrows():
                m = row["month"].month
                seasonality[m] = row["staff_cost"] / h2_avg if h2_avg > 0 else 1.0

        current_year_data = monthly[
            monthly["month_str"].str.startswith(str(current_year))
        ].copy()
        forecast_months = []
        current_month = pd.Timestamp.now().month
        for m in range(current_month + 1, 13):
            factor = seasonality.get(m, 1.0)
            forecast_months.append({
                "month_str": f"{current_year}-{m:02d}",
                "staff_cost": avg_cost * factor,
                "revenue": avg_rev * factor,
                "is_forecast": True,
            })
        current_year_data["is_forecast"] = False
        forecast_df = pd.concat(
            [current_year_data[["month_str", "staff_cost", "revenue"]].assign(is_forecast=False),
             pd.DataFrame(forecast_months)],
            ignore_index=True,
        )

        fig_fc = go.Figure()
        actuals = forecast_df[~forecast_df["is_forecast"]]
        forecasts = forecast_df[forecast_df["is_forecast"]]

        fig_fc.add_trace(go.Bar(
            x=actuals["month_str"], y=actuals["revenue"],
            name="Revenue (Actual)", marker_color=COLORS["revenue"],
            hovertemplate="EUR %{y:,.0f}<extra></extra>",
        ))
        fig_fc.add_trace(go.Bar(
            x=forecasts["month_str"], y=forecasts["revenue"],
            name="Revenue (Forecast)", marker_color=COLORS["revenue"],
            opacity=0.4, marker_pattern_shape="/",
            hovertemplate="EUR %{y:,.0f} (forecast)<extra></extra>",
        ))
        fig_fc.add_trace(go.Bar(
            x=actuals["month_str"], y=actuals["staff_cost"],
            name="Staff Cost (Actual)", marker_color=COLORS["cost"],
            hovertemplate="EUR %{y:,.0f}<extra></extra>",
        ))
        fig_fc.add_trace(go.Bar(
            x=forecasts["month_str"], y=forecasts["staff_cost"],
            name="Staff Cost (Forecast)", marker_color=COLORS["cost"],
            opacity=0.4, marker_pattern_shape="/",
            hovertemplate="EUR %{y:,.0f} (forecast)<extra></extra>",
        ))
        fig_fc.update_layout(
            **CHART_LAYOUT,
            title=f"EOY Forecast {current_year} — Actuals + Forecast",
            barmode="group",
            yaxis_title="EUR",
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        charts["financial_eoy_forecast"] = fig_fc

    return charts


# ============================================================
# 2. PEOPLE REPORT
# ============================================================

def report_people(te: pd.DataFrame) -> pd.DataFrame:
    """
    Billable Utilisation, Avg Staff Cost/Hour, Total Hours per person.

    Utilisation = billable_hours / total_tracked_hours (everyone logs all time)
    Avg Cost/Hour = work_cost / hours (owners overridden to EUR 15K/month)
    """
    print("\n" + "=" * 70)
    print("PEOPLE REPORT")
    print("=" * 70)
    print(f"\n  {ANNOTATIONS['utilisation']}")
    print(f"  {ANNOTATIONS['staff_cost']}\n")

    people = te.groupby(["person_name", "team"]).agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
        entries=("id", "count"),
    ).reset_index()

    people["utilisation_pct"] = (
        people["billable_hours"] / people["total_hours"] * 100
    ).round(1)
    people["avg_cost_per_hour"] = (
        people["staff_cost"] / people["total_hours"]
    ).round(2)

    people = people.sort_values("total_hours", ascending=False)

    print(f"  {'Person':<30} {'Team':<22} {'Hours':>7} {'Bill.':>7} {'Util%':>6} {'EUR/h':>7} {'Staff Cost':>12}")
    print(f"  {'-'*30} {'-'*22} {'-'*7} {'-'*7} {'-'*6} {'-'*7} {'-'*12}")
    for _, r in people.iterrows():
        print(
            f"  {r['person_name']:<30} {r['team']:<22} "
            f"{r['total_hours']:>7,.1f} {r['billable_hours']:>7,.1f} "
            f"{r['utilisation_pct']:>5.1f}% {r['avg_cost_per_hour']:>7,.2f} "
            f"EUR {r['staff_cost']:>9,.0f}"
        )

    # Team summary
    print(f"\n  Team Summary:")
    team_summary = te.groupby("team").agg(
        hours=("hours", "sum"),
        billable=("billable_hours", "sum"),
        cost=("work_cost", "sum"),
        people=("person_name", "nunique"),
    ).reset_index()
    team_summary["util_pct"] = (team_summary["billable"] / team_summary["hours"] * 100).round(1)
    team_summary["avg_rate"] = (team_summary["cost"] / team_summary["hours"]).round(2)

    print(f"  {'Team':<28} {'People':>6} {'Hours':>8} {'Util%':>6} {'EUR/h':>7} {'Staff Cost':>12}")
    print(f"  {'-'*28} {'-'*6} {'-'*8} {'-'*6} {'-'*7} {'-'*12}")
    for _, r in team_summary.iterrows():
        print(
            f"  {r['team']:<28} {r['people']:>6} {r['hours']:>8,.1f} "
            f"{r['util_pct']:>5.1f}% {r['avg_rate']:>7,.2f} EUR {r['cost']:>9,.0f}"
        )

    _save_report(people, "people")

    # --- Charts ---
    charts = _charts_people(people, team_summary)

    return people, charts


def _charts_people(people: pd.DataFrame, team_summary: pd.DataFrame) -> dict[str, go.Figure]:
    """Build people report charts. Returns {name: figure}."""
    charts = {}
    # 1. Billable Utilisation % — horizontal bar, ranked, color-coded
    df = people.sort_values("utilisation_pct", ascending=True).copy()
    fig_util = px.bar(
        df,
        x="utilisation_pct",
        y="person_name",
        orientation="h",
        color="utilisation_pct",
        color_continuous_scale=[COLORS["overbudget"], COLORS["warning"], COLORS["on_track"]],
        range_color=[0, 100],
        labels={"utilisation_pct": "Utilisation %", "person_name": ""},
        hover_data={"total_hours": ":.1f", "billable_hours": ":.1f", "team": True},
    )
    fig_util.update_layout(
        **CHART_LAYOUT,
        title="Billable Utilisation by Person",
        height=max(400, len(df) * 25),
        coloraxis_colorbar_title="Util %",
    )
    charts["people_utilisation"] = fig_util

    # 2. Total Hours per Person — stacked horizontal bar (billable + non-billable)
    df2 = people.sort_values("total_hours", ascending=True).copy()
    df2["non_billable_hours"] = df2["total_hours"] - df2["billable_hours"]

    fig_hours = go.Figure()
    fig_hours.add_trace(go.Bar(
        y=df2["person_name"], x=df2["billable_hours"],
        name="Billable", orientation="h",
        marker_color=COLORS["billable"],
        hovertemplate="%{y}: %{x:,.1f}h billable<extra></extra>",
    ))
    fig_hours.add_trace(go.Bar(
        y=df2["person_name"], x=df2["non_billable_hours"],
        name="Non-Billable", orientation="h",
        marker_color=COLORS["non_billable"],
        hovertemplate="%{y}: %{x:,.1f}h non-billable<extra></extra>",
    ))
    fig_hours.update_layout(
        **CHART_LAYOUT,
        title="Total Hours per Person (Billable + Non-Billable)",
        barmode="stack",
        xaxis_title="Hours",
        height=max(400, len(df2) * 25),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
    )
    charts["people_hours"] = fig_hours

    # 3. Team Summary — grouped bar chart
    fig_team = go.Figure()
    fig_team.add_trace(go.Bar(
        x=team_summary["team"], y=team_summary["hours"],
        name="Total Hours", marker_color=COLORS["non_billable"],
        hovertemplate="%{x}: %{y:,.0f}h total<extra></extra>",
    ))
    fig_team.add_trace(go.Bar(
        x=team_summary["team"], y=team_summary["billable"],
        name="Billable Hours", marker_color=COLORS["billable"],
        hovertemplate="%{x}: %{y:,.0f}h billable<extra></extra>",
    ))
    fig_team.update_layout(
        **CHART_LAYOUT,
        title="Team Summary — Hours",
        barmode="group",
        yaxis_title="Hours",
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
    )
    charts["people_team_summary"] = fig_team

    return charts


# ============================================================
# 3. PROJECT REPORT
# ============================================================

def report_project(deals: pd.DataFrame, te: pd.DataFrame) -> pd.DataFrame:
    """
    Overbudget flags (hours + EUR), ACPH + ARPH by service type.

    Overbudget: Green <70%, Amber 70-100%, Red >100% (dual hours + EUR)
    ACPH = staff_cost / worked_hours by service type
    ARPH = revenue / billable_hours by service type
    """
    print("\n" + "=" * 70)
    print("PROJECT REPORT")
    print("=" * 70)
    print(f"\n  {ANNOTATIONS['overbudget']}")
    print(f"  {ANNOTATIONS['acph']}\n")

    # --- Overbudget flags ---
    client_budgets = deals[
        (deals["budget"] == True) & (deals["deal_type_id"] == 2)
    ].copy()

    # Only deals with budgeted hours or budget_total
    flaggable = client_budgets[
        (client_budgets["budgeted_hours"] > 0) | (client_budgets["budget_total"] > 0)
    ].copy()

    def flag_deal(row):
        hours_pct = (
            row["worked_hours"] / row["budgeted_hours"] * 100
            if row["budgeted_hours"] > 0 else 0
        )
        cost_pct = (
            row["cost"] / row["budget_total"] * 100
            if row["budget_total"] > 0 else 0
        )
        if hours_pct > 100 or cost_pct > 100:
            return "RED"
        elif hours_pct > 70 or cost_pct > 70:
            return "AMBER"
        return "GREEN"

    flaggable["flag"] = flaggable.apply(flag_deal, axis=1)
    flaggable["hours_burn_pct"] = (
        flaggable["worked_hours"] / flaggable["budgeted_hours"].replace(0, float("nan")) * 100
    ).round(1)
    flaggable["cost_burn_pct"] = (
        flaggable["cost"] / flaggable["budget_total"].replace(0, float("nan")) * 100
    ).round(1)

    flag_counts = flaggable["flag"].value_counts()
    print(f"  Overbudget Summary: {flag_counts.get('GREEN', 0)} Green, "
          f"{flag_counts.get('AMBER', 0)} Amber, {flag_counts.get('RED', 0)} Red")
    print(f"  ({len(flaggable)} deals with budget data)\n")

    # Show RED deals
    red_deals = flaggable[flaggable["flag"] == "RED"].sort_values("hours_burn_pct", ascending=False)
    if not red_deals.empty:
        print(f"  RED Deals (top 15):")
        print(f"  {'Deal':<45} {'Hours%':>7} {'EUR%':>7} {'Worked':>8} {'Budget':>8}")
        print(f"  {'-'*45} {'-'*7} {'-'*7} {'-'*8} {'-'*8}")
        for _, r in red_deals.head(15).iterrows():
            name = str(r.get("name", "?"))[:44]
            print(
                f"  {name:<45} {r['hours_burn_pct']:>6.0f}% {r['cost_burn_pct']:>6.0f}% "
                f"{r['worked_hours']:>7.0f}h {r['budgeted_hours']:>7.0f}h"
            )

    # --- ACPH + ARPH by service type ---
    print(f"\n  Cost & Revenue Per Hour by Service Type:")

    stype_metrics = te.groupby("service_type").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
    ).reset_index()

    # Get revenue per service type from services snapshot (or approximate)
    # For now, compute from time entries where we know billable hours
    stype_metrics["acph"] = (
        stype_metrics["staff_cost"] / stype_metrics["total_hours"]
    ).round(2)

    # ARPH from per-entry revenue (billable_time × service_price)
    if "entry_revenue" in te.columns:
        stype_revenue = te.groupby("service_type")["entry_revenue"].sum().reset_index()
        stype_revenue.columns = ["service_type", "allocated_revenue"]
        stype_metrics = stype_metrics.merge(stype_revenue, on="service_type", how="left")
        stype_metrics["allocated_revenue"] = stype_metrics["allocated_revenue"].fillna(0)
        stype_metrics["arph"] = (
            stype_metrics["allocated_revenue"] / stype_metrics["billable_hours"].replace(0, float("nan"))
        ).round(2)
    else:
        stype_metrics["allocated_revenue"] = 0
        stype_metrics["arph"] = 0

    stype_metrics["margin_per_hour"] = (stype_metrics["arph"] - stype_metrics["acph"]).round(2)
    stype_metrics = stype_metrics.sort_values("total_hours", ascending=False)

    print(f"  {'Service Type':<45} {'Hours':>7} {'ACPH':>8} {'ARPH':>8} {'Gap':>8}")
    print(f"  {'-'*45} {'-'*7} {'-'*8} {'-'*8} {'-'*8}")
    for _, r in stype_metrics.iterrows():
        name = str(r["service_type"] or "(no type)")[:44]
        arph = f"{r['arph']:>7.2f}" if pd.notna(r["arph"]) else "    N/A"
        gap = f"{r['margin_per_hour']:>7.2f}" if pd.notna(r["margin_per_hour"]) else "    N/A"
        print(
            f"  {name:<45} {r['total_hours']:>7,.0f} "
            f"{r['acph']:>7.2f} {arph} {gap}"
        )

    _save_report(flaggable[["name", "flag", "hours_burn_pct", "cost_burn_pct",
                             "worked_hours", "budgeted_hours", "revenue", "cost"]]
                 if not flaggable.empty else pd.DataFrame(),
                 "project_overbudget")
    _save_report(stype_metrics, "project_acph_arph")

    # --- Charts ---
    charts = _charts_project(flaggable, stype_metrics)

    return flaggable, charts


def _charts_project(flaggable: pd.DataFrame, stype_metrics: pd.DataFrame) -> dict[str, go.Figure]:
    """Build project report charts. Returns {name: figure}."""
    charts = {}
    # 1. ACPH vs ARPH by Service Type — grouped bar
    df = stype_metrics[stype_metrics["total_hours"] > 10].copy()
    df["service_type"] = df["service_type"].fillna("(no type)")
    df = df.sort_values("total_hours", ascending=False)

    fig_rates = go.Figure()
    fig_rates.add_trace(go.Bar(
        x=df["service_type"], y=df["acph"],
        name="ACPH (Cost)", marker_color=COLORS["cost"],
        hovertemplate="%{x}<br>ACPH: EUR %{y:.2f}/h<extra></extra>",
    ))
    fig_rates.add_trace(go.Bar(
        x=df["service_type"], y=df["arph"],
        name="ARPH (Revenue)", marker_color=COLORS["revenue"],
        hovertemplate="%{x}<br>ARPH: EUR %{y:.2f}/h<extra></extra>",
    ))
    fig_rates.update_layout(
        **CHART_LAYOUT,
        title="ACPH vs ARPH by Service Type",
        barmode="group",
        yaxis_title="EUR / Hour",
        xaxis_tickangle=-45,
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
    )
    charts["project_acph_arph"] = fig_rates

    # 2. Overbudget Distribution — stacked bar showing counts
    if not flaggable.empty:
        flag_counts = flaggable["flag"].value_counts().reindex(
            ["GREEN", "AMBER", "RED"], fill_value=0
        )
        flag_colors = {
            "GREEN": COLORS["on_track"],
            "AMBER": COLORS["warning"],
            "RED": COLORS["overbudget"],
        }
        fig_ob = go.Figure()
        for flag_name in ["GREEN", "AMBER", "RED"]:
            fig_ob.add_trace(go.Bar(
                x=[flag_name],
                y=[flag_counts.get(flag_name, 0)],
                name=flag_name,
                marker_color=flag_colors[flag_name],
                hovertemplate=f"{flag_name}: %{{y}} deals<extra></extra>",
            ))
        fig_ob.update_layout(
            **CHART_LAYOUT,
            title="Overbudget Distribution",
            yaxis_title="Number of Deals",
            showlegend=False,
        )
        charts["project_overbudget_dist"] = fig_ob

    return charts


# ============================================================
# 4. CLIENT REPORT
# ============================================================

def report_client(te: pd.DataFrame, deals: pd.DataFrame) -> pd.DataFrame:
    """
    Revenue, Staff Cost, Gross Margin, Overbudget count, ACPH, ARPH per client.
    """
    print("\n" + "=" * 70)
    print("CLIENT REPORT")
    print("=" * 70)
    print(f"\n  {ANNOTATIONS['revenue']}")
    print(f"  {ANNOTATIONS['staff_cost']}\n")

    # Revenue from deals grouped by client
    client_budgets = deals[
        (deals["budget"] == True) & (deals["deal_type_id"] == 2)
    ].copy()

    deal_financials = client_budgets.groupby("company_name").agg(
        revenue=("revenue", "sum"),
        deal_cost=("cost", "sum"),
        deal_count=("id", "count"),
        budgeted_hours=("budgeted_hours", "sum"),
        worked_hours_deal=("worked_hours", "sum"),
    ).reset_index()
    deal_financials.rename(columns={"company_name": "client_name"}, inplace=True)

    # Hours + cost from time entries grouped by client
    te_client = te[te["client_name"].notna()].copy()
    client_te = te_client.groupby("client_name").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
        people=("person_name", "nunique"),
    ).reset_index()

    # Merge
    clients = client_te.merge(deal_financials, on="client_name", how="left")
    clients["revenue"] = clients["revenue"].fillna(0)
    clients["gross_margin"] = clients["revenue"] - clients["staff_cost"]
    clients["margin_pct"] = (
        clients["gross_margin"] / clients["revenue"].replace(0, float("nan")) * 100
    ).round(1)
    clients["acph"] = (clients["staff_cost"] / clients["total_hours"]).round(2)
    clients["arph"] = (
        clients["revenue"] / clients["billable_hours"].replace(0, float("nan"))
    ).round(2)

    # Overbudget count per client
    def count_flagged(client_name):
        cd = client_budgets[client_budgets["company_name"] == client_name]
        flagged = 0
        for _, row in cd.iterrows():
            hp = row["worked_hours"] / row["budgeted_hours"] * 100 if row["budgeted_hours"] > 0 else 0
            cp = row["cost"] / row["budget_total"] * 100 if row["budget_total"] > 0 else 0
            if hp > 70 or cp > 70:
                flagged += 1
        return flagged

    clients["overbudget_deals"] = clients["client_name"].apply(count_flagged)

    clients = clients.sort_values("revenue", ascending=False)

    # Print top clients
    print(f"  {'Client':<40} {'Revenue':>12} {'Staff Cost':>12} {'Margin':>8} {'ACPH':>7} {'ARPH':>7} {'OB':>3}")
    print(f"  {'-'*40} {'-'*12} {'-'*12} {'-'*8} {'-'*7} {'-'*7} {'-'*3}")
    for _, r in clients.head(25).iterrows():
        name = str(r["client_name"])[:39]
        margin = f"{r['margin_pct']:>6.1f}%" if pd.notna(r["margin_pct"]) else "    N/A"
        arph = f"{r['arph']:>7.2f}" if pd.notna(r["arph"]) else "    N/A"
        ob = int(r["overbudget_deals"]) if pd.notna(r["overbudget_deals"]) else 0
        print(
            f"  {name:<40} EUR {r['revenue']:>8,.0f} EUR {r['staff_cost']:>8,.0f} "
            f"{margin} {r['acph']:>7.2f} {arph} {ob:>3}"
        )

    print(f"\n  Total clients: {len(clients)}")
    print(f"  Clients with overbudget deals: {(clients['overbudget_deals'] > 0).sum()}")

    _save_report(clients, "client")

    # --- Charts ---
    charts = _charts_client(clients)

    return clients, charts


def _charts_client(clients: pd.DataFrame) -> dict[str, go.Figure]:
    """Build client report charts. Returns {name: figure}."""
    charts = {}
    # 1. Top 15 Clients by Revenue — horizontal bar with margin % annotation
    top15 = clients.head(15).sort_values("revenue", ascending=True).copy()

    fig_rev = go.Figure()
    fig_rev.add_trace(go.Bar(
        y=top15["client_name"], x=top15["revenue"],
        orientation="h", marker_color=COLORS["revenue"],
        hovertemplate=(
            "%{y}<br>Revenue: EUR %{x:,.0f}<br>"
            "Staff Cost: EUR %{customdata[0]:,.0f}<br>"
            "Margin: %{customdata[1]:.1f}%<extra></extra>"
        ),
        customdata=top15[["staff_cost", "margin_pct"]].values,
    ))
    # Add margin % annotations
    for _, r in top15.iterrows():
        if pd.notna(r["margin_pct"]):
            fig_rev.add_annotation(
                x=r["revenue"],
                y=r["client_name"],
                text=f" {r['margin_pct']:.0f}%",
                showarrow=False,
                xanchor="left",
                font=dict(
                    size=10,
                    color=COLORS["profit_pos"] if r["margin_pct"] >= 0 else COLORS["profit_neg"],
                ),
            )
    fig_rev.update_layout(
        **CHART_LAYOUT,
        title="Top 15 Clients by Revenue",
        xaxis_title="Revenue (EUR)",
        height=500,
    )
    charts["client_top15_revenue"] = fig_rev

    # 2. Client Profitability — horizontal bar sorted by margin %, color-coded
    profitable = clients[clients["revenue"] > 0].copy()
    profitable = profitable.sort_values("margin_pct", ascending=True)
    # Color-code: green for positive, red for negative
    profitable["color"] = profitable["margin_pct"].apply(
        lambda x: COLORS["profit_pos"] if x >= 0 else COLORS["profit_neg"]
    )

    fig_profit = go.Figure()
    fig_profit.add_trace(go.Bar(
        y=profitable["client_name"],
        x=profitable["margin_pct"],
        orientation="h",
        marker_color=profitable["color"],
        hovertemplate=(
            "%{y}<br>Margin: %{x:.1f}%<br>"
            "Revenue: EUR %{customdata[0]:,.0f}<br>"
            "Gross Margin: EUR %{customdata[1]:,.0f}<extra></extra>"
        ),
        customdata=profitable[["revenue", "gross_margin"]].values,
    ))
    fig_profit.add_vline(x=0, line_dash="dash", line_color="gray")
    fig_profit.update_layout(
        **CHART_LAYOUT,
        title="Client Profitability (Gross Margin %)",
        xaxis_title="Gross Margin %",
        height=max(400, len(profitable) * 22),
    )
    charts["client_profitability"] = fig_profit

    return charts


# ============================================================
# 5. HYGIENE REPORT
# ============================================================

def report_hygiene(te: pd.DataFrame) -> pd.DataFrame:
    """
    Full hygiene suite: missing notes, zero-hour entries, unapproved,
    overspend tracking, missing service type, per-person breakdown.
    """
    print("\n" + "=" * 70)
    print("HYGIENE REPORT")
    print("=" * 70)

    total = len(te)

    # --- Missing notes ---
    missing_notes = (~te["has_note"]).sum()
    missing_notes_pct = missing_notes / total * 100
    print(f"\n  Missing Notes: {missing_notes:,} / {total:,} ({missing_notes_pct:.1f}%)")

    monthly_notes = te.groupby("month").agg(
        total=("id", "count"),
        with_note=("has_note", "sum"),
    ).reset_index()
    monthly_notes["pct_with_note"] = (monthly_notes["with_note"] / monthly_notes["total"] * 100).round(1)

    print(f"\n  Monthly Note Compliance:")
    for _, r in monthly_notes.iterrows():
        bar = "#" * int(r["pct_with_note"] / 2)
        print(f"    {r['month']}: {r['pct_with_note']:>5.1f}% {bar}")

    # --- Zero-hour entries ---
    zero_hours = (te["hours"] == 0).sum()
    print(f"\n  Zero-Hour Entries: {zero_hours:,} ({zero_hours / total * 100:.1f}%)")

    # --- Unapproved entries ---
    if "approved" in te.columns:
        unapproved = (~te["approved"].astype(bool)).sum()
        print(f"  Unapproved Entries: {unapproved:,} ({unapproved / total * 100:.1f}%)")

    # --- Overspend tracking ---
    if "billing_type_id" in te.columns:
        non_billable = te[te["billing_type_id"] == 3]
        overspend_keywords = ["overspend", "overrun", "over budget", "extra"]
        overspend_entries = non_billable[
            non_billable["service_name"].str.lower().str.contains(
                "|".join(overspend_keywords), na=False
            )
        ]
        overspend_hours = overspend_entries["hours"].sum()
        overspend_cost = overspend_entries["work_cost"].sum()
        print(f"\n  Overspend Entries: {len(overspend_entries):,} entries, "
              f"{overspend_hours:,.1f}h, EUR {overspend_cost:,.0f} staff cost")
        if not overspend_entries.empty:
            print(f"  Overspend by deal:")
            by_deal = overspend_entries.groupby("deal_name").agg(
                hours=("hours", "sum"), cost=("work_cost", "sum")
            ).sort_values("hours", ascending=False).head(10)
            for deal, r in by_deal.iterrows():
                print(f"    {deal}: {r['hours']:.1f}h, EUR {r['cost']:,.0f}")

    # --- Missing service type ---
    if "service_type" in te.columns:
        no_stype = te["service_type"].isna().sum()
        print(f"\n  Missing Service Type: {no_stype:,} entries ({no_stype / total * 100:.1f}%)")

    # --- Per-person breakdown ---
    print(f"\n  Per-Person Hygiene:")
    person_hygiene = te.groupby(["person_name", "team"]).agg(
        total=("id", "count"),
        missing_notes=("has_note", lambda x: (~x).sum()),
        zero_hours=("hours", lambda x: (x == 0).sum()),
    ).reset_index()

    if "approved" in te.columns:
        unapproved_by_person = te.groupby("person_name")["approved"].apply(
            lambda x: (~x.astype(bool)).sum()
        ).reset_index()
        unapproved_by_person.columns = ["person_name", "unapproved"]
        person_hygiene = person_hygiene.merge(unapproved_by_person, on="person_name", how="left")
    else:
        person_hygiene["unapproved"] = 0

    person_hygiene["note_pct"] = (
        (person_hygiene["total"] - person_hygiene["missing_notes"]) / person_hygiene["total"] * 100
    ).round(1)
    person_hygiene = person_hygiene.sort_values("note_pct", ascending=True)

    print(f"  {'Person':<30} {'Team':<22} {'Note%':>6} {'Zero':>5} {'Unapp':>6}")
    print(f"  {'-'*30} {'-'*22} {'-'*6} {'-'*5} {'-'*6}")
    for _, r in person_hygiene.iterrows():
        print(
            f"  {r['person_name']:<30} {r['team']:<22} "
            f"{r['note_pct']:>5.1f}% {int(r['zero_hours']):>5} {int(r['unapproved']):>6}"
        )

    _save_report(person_hygiene, "hygiene_by_person")
    _save_report(monthly_notes, "hygiene_monthly_notes")

    # --- Charts ---
    charts = _charts_hygiene(monthly_notes, person_hygiene)

    return person_hygiene, charts


def _charts_hygiene(monthly_notes: pd.DataFrame, person_hygiene: pd.DataFrame) -> dict[str, go.Figure]:
    """Build hygiene report charts. Returns {name: figure}."""
    charts = {}
    # 1. Note Compliance Trend — line chart over months
    mn = monthly_notes.copy()
    mn["month_str"] = mn["month"].astype(str)

    fig_trend = go.Figure()
    fig_trend.add_trace(go.Scatter(
        x=mn["month_str"], y=mn["pct_with_note"],
        mode="lines+markers",
        line=dict(color=COLORS["revenue"], width=2),
        marker=dict(size=8),
        hovertemplate="%{x}<br>Note compliance: %{y:.1f}%<br>Entries: %{customdata}<extra></extra>",
        customdata=mn["total"],
    ))
    fig_trend.add_hline(y=100, line_dash="dash", line_color=COLORS["on_track"], opacity=0.5)
    fig_trend.update_layout(
        **CHART_LAYOUT,
        title="Note Compliance Trend (% of entries with notes)",
        yaxis_title="% With Note",
        yaxis_range=[0, 105],
    )
    charts["hygiene_note_trend"] = fig_trend

    # 2. Per-Person Note Compliance — horizontal bar, ranked
    df = person_hygiene.sort_values("note_pct", ascending=True).copy()
    # Color by compliance level
    df["color"] = df["note_pct"].apply(
        lambda x: COLORS["on_track"] if x >= 80
        else COLORS["warning"] if x >= 50
        else COLORS["overbudget"]
    )

    fig_person = go.Figure()
    fig_person.add_trace(go.Bar(
        y=df["person_name"], x=df["note_pct"],
        orientation="h",
        marker_color=df["color"],
        hovertemplate=(
            "%{y}<br>Note compliance: %{x:.1f}%<br>"
            "Missing: %{customdata[0]}<br>"
            "Team: %{customdata[1]}<extra></extra>"
        ),
        customdata=df[["missing_notes", "team"]].values,
    ))
    fig_person.add_vline(x=80, line_dash="dash", line_color=COLORS["on_track"], opacity=0.5,
                         annotation_text="80% target")
    fig_person.update_layout(
        **CHART_LAYOUT,
        title="Per-Person Note Compliance",
        xaxis_title="% Entries With Note",
        xaxis_range=[0, 105],
        height=max(400, len(df) * 25),
    )
    charts["hygiene_person_notes"] = fig_person

    return charts


# ============================================================
# DASHBOARD DATA EXPORT
# ============================================================

def _compute_rate_card_actions(services_df: pd.DataFrame | None) -> list[dict]:
    """Find T&M services below €125/h that could be updated."""
    if services_df is None:
        return []
    tm = services_df[
        (services_df["billing_type_id"] == 2)
        & (services_df["billable"] == True)
        & (~services_df["deal_name"].str.contains("PSO", case=False, na=False))
    ].copy()
    tm["price_eur"] = pd.to_numeric(tm["price"], errors="coerce") / 100.0
    tm["worked_hrs"] = pd.to_numeric(tm["worked_time"], errors="coerce") / 60.0
    below = tm[(tm["price_eur"] > 0) & (tm["price_eur"] < 125) & (tm["worked_hrs"] > 0)]

    groups = below.groupby("price_eur").agg(
        count=("id", "count"),
        total_worked=("worked_hrs", "sum"),
    ).reset_index()
    groups["uplift"] = (125 - groups["price_eur"]) * groups["total_worked"]

    actions = []
    for _, r in groups.sort_values("uplift", ascending=False).iterrows():
        actions.append({
            "current_rate": round(r["price_eur"], 0),
            "target_rate": 125,
            "service_count": int(r["count"]),
            "worked_hours": round(r["total_worked"], 0),
            "uplift": round(r["uplift"], 0),
        })
    return actions


def get_dashboard_data(te: pd.DataFrame, deals: pd.DataFrame) -> dict[str, list[dict]]:
    """Compute all aggregated data needed for the interactive dashboard.

    Returns a dict of {name: list_of_row_dicts} ready for JSON serialization.
    Each row dict has only JSON-safe types (str, int, float, None).
    """

    # --- Financial monthly (revenue from API, cost from time entries) ---
    monthly_te = te.groupby("month").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
    ).reset_index()
    monthly_te["month_str"] = monthly_te["month"].astype(str)

    client_budgets = deals[
        (deals["budget"] == True) & (deals["deal_type_id"] == 2)
    ].copy()

    # --- Contingent detection (pre-invoiced hour banks) ---
    # Source of truth: the contingenten Excel file (2026 tab, "Remaining" row).
    # Each Excel column maps to a Productive company via keyword matching.
    # Checked in order; first match wins. More specific patterns first.
    _CONTINGENT_ENTRY_MAP = [
        # (keyword_in_excel_header, productive_company_name)
        # --- CGK / Cegeka (specific before generic) ---
        ("cgk groep", "CEGEKA GROEP NV (CGK)"),
        ("cgk modules", "CEGEKA GROEP NV (CGK)"),
        ("cgk (modules", "Cegeka (CGK) nv"),
        ("ctg lu", "CTG IT Solutions"),
        ("ctg", "CTG"),
        ("cegeka 2025 business", "Cegeka Business Solutions België NV (CBS EU)"),
        ("cegeka 2025 trust", "Cegeka (CGK) nv"),
        ("cegeka 2025 nl", "Cegeka (CGK) nv"),
        # --- Other clients ---
        ("agoria", None),  # not active in current Productive data
        ("beaulieu", "Beaulieu International Group NV"),
        ("cafca", "Cafca nv"),
        ("company web", "Companyweb bv"),
        ("element61", "element61 N.V."),
        ("flenhealth", "Flen Health SA - Luxembourg"),
        ("goed", "Goed"),
        ("leasebroker", "Lease-Broker"),
        ("liantis", "Liantis corporate vzw"),
        ("modular", "PITS NV - Modular Lighting Instruments"),
        ("nexuzhealth", "Nexuzhealth"),
        ("pollergroup", "Pollet Water Group"),
        ("symeta", None),  # not active in current Productive data
        ("vreg", "Vlaamse Nutsregulator"),
        ("atradius be", "Atradius Crédito y Caución S.A. de Seguros y Reaseguros (BE)"),
        ("atradius fr", "Atradius Crédito y Caución S.A. de Seguros y Reaseguros (FR)"),
        ("atradius group", "Atradius Crédito y Caución S.A. de Seguros y Reaseguros (Group)"),
        ("atradius", "Atradius Crédito y Caución S.A. de Seguros y Reaseguros (Group)"),
        ("cambridge", "Cambridge Design Partners"),
    ]

    # Read remaining balances from the contingenten Excel
    from collections import defaultdict
    _contingent_remaining = defaultdict(float)  # company_name → remaining €
    _contingent_company_set = set()
    _CONTINGENT_EXCEL = os.path.join(
        os.path.expanduser("~"),
        "Library/CloudStorage/OneDrive-Gedeeldebibliotheken-T-WESSC°BV/"
        "Storage-Meja - Documenten/Meja/Always On Group/Leadstreet/"
        "2. Finance/2. Reporting/2026/Contingenten/"
        "260326 2026-2025-2024-2023-2022-2021-contingenten.xlsx"
    )
    try:
        import openpyxl
        _wb = openpyxl.load_workbook(_CONTINGENT_EXCEL, data_only=True)
        _ws = _wb["2026"]
        _headers = list(list(_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0])
        _rem_row = None
        for _row in _ws.iter_rows(min_row=3, max_row=6, values_only=True):
            if list(_row)[1] == "Remaining":
                _rem_row = list(_row)
                break
        if _rem_row:
            for _i, _h in enumerate(_headers):
                if not _h or str(_h).strip() in ("", "2026") or _i <= 1:
                    continue
                _rem = _rem_row[_i] if _i < len(_rem_row) and isinstance(_rem_row[_i], (int, float)) else 0
                _h_lower = str(_h).strip().lower()
                _matched_company = None
                for _kw, _comp in _CONTINGENT_ENTRY_MAP:
                    if _kw in _h_lower:
                        _matched_company = _comp
                        break
                if _matched_company:
                    _contingent_remaining[_matched_company] += _rem
                    _contingent_company_set.add(_matched_company)
                elif _matched_company is None and any(_kw in _h_lower for _kw, _ in _CONTINGENT_ENTRY_MAP):
                    pass  # known client, no active Productive match (Agoria, Symeta)
                else:
                    print(f"  WARNING: Contingent entry '{str(_h).strip()[:50]}' has no mapping")
        _wb.close()
        print(f"\n  CONTINGENT REMAINING (from Excel):")
        for _comp, _rem in sorted(_contingent_remaining.items(), key=lambda x: -x[1]):
            print(f"    {_comp[:50]:50s}  €{_rem:>10,.2f}")
    except FileNotFoundError:
        print(f"\n  WARNING: Contingenten Excel not found, contingent detection disabled")
    except Exception as e:
        print(f"\n  WARNING: Error reading contingenten Excel: {e}")

    # Flag contingent deals and set excess from Excel remaining values
    cb = client_budgets
    cb["is_contingent"] = cb["company_name"].isin(_contingent_company_set)
    cb["contingent_excess"] = 0.0
    cb["adjusted_revenue"] = cb["revenue"]
    client_budgets = cb

    # Revenue from Productive's financial_item_reports API (recognized + invoiced)
    fir = load_financial_item_reports()
    if fir is not None:
        monthly = monthly_te.merge(
            fir[["month_str", "revenue", "invoiced"]], on="month_str", how="left"
        )
        monthly["revenue"] = monthly["revenue"].fillna(0)
        monthly["invoiced"] = monthly["invoiced"].fillna(0)
    else:
        print("  WARNING: No financial_item_reports snapshot found, revenue will be 0")
        monthly = monthly_te.copy()
        monthly["revenue"] = 0
        monthly["invoiced"] = 0

    monthly["gross_margin"] = monthly["revenue"] - monthly["staff_cost"]
    monthly["margin_pct"] = (
        monthly["gross_margin"] / monthly["revenue"].replace(0, float("nan")) * 100
    ).fillna(0).round(1)
    monthly["util_pct"] = (
        monthly["billable_hours"] / monthly["total_hours"] * 100
    ).fillna(0).round(1)

    # Invoiced margin
    monthly["inv_gross_margin"] = monthly["invoiced"] - monthly["staff_cost"]
    monthly["inv_margin_pct"] = (
        monthly["inv_gross_margin"] / monthly["invoiced"].replace(0, float("nan")) * 100
    ).fillna(0).round(1)

    financial_monthly = monthly[
        ["month_str", "revenue", "invoiced", "staff_cost",
         "gross_margin", "margin_pct", "inv_gross_margin", "inv_margin_pct",
         "total_hours", "billable_hours", "util_pct"]
    ].rename(columns={"month_str": "month"})

    # --- People monthly (person × month) ---
    agg_dict = {
        "hours": ("hours", "sum"),
        "billable_hours": ("billable_hours", "sum"),
        "staff_cost": ("work_cost", "sum"),
    }
    if "entry_revenue" in te.columns:
        agg_dict["entry_revenue"] = ("entry_revenue", "sum")
    people_monthly = te.groupby(["person_name", "person_id", "team", "month"]).agg(**agg_dict).reset_index()
    people_monthly["month"] = people_monthly["month"].astype(str)
    if "entry_revenue" not in people_monthly.columns:
        people_monthly["entry_revenue"] = 0
    people_monthly["utilisation_pct"] = (
        people_monthly["billable_hours"] / people_monthly["hours"] * 100
    ).fillna(0).round(1)
    # Revenue will be enriched with recognized revenue later (after fir_budget is loaded)

    # --- Team summary ---
    team_summary = te.groupby("team").agg(
        people=("person_name", "nunique"),
        hours=("hours", "sum"),
        billable=("billable_hours", "sum"),
        cost=("work_cost", "sum"),
    ).reset_index()
    team_summary["util_pct"] = (
        team_summary["billable"] / team_summary["hours"] * 100
    ).fillna(0).round(1)

    # --- Service type metrics ---
    stype = te.groupby("service_type").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
    ).reset_index()
    stype["acph"] = (stype["staff_cost"] / stype["total_hours"]).round(2)
    # ARPH from per-entry revenue (billable_time × service_price)
    if "entry_revenue" in te.columns:
        stype_rev = te.groupby("service_type")["entry_revenue"].sum().reset_index()
        stype_rev.columns = ["service_type", "allocated_revenue"]
        stype = stype.merge(stype_rev, on="service_type", how="left")
        stype["allocated_revenue"] = stype["allocated_revenue"].fillna(0)
        stype["arph"] = (
            stype["allocated_revenue"] / stype["billable_hours"].replace(0, float("nan"))
        ).fillna(0).round(2)
    else:
        stype["arph"] = 0
    stype["margin_per_hour"] = (stype["arph"] - stype["acph"]).fillna(0).round(2)
    stype["service_type"] = stype["service_type"].fillna("(no type)")

    # --- Per-deal summary (from time entries + per-budget recognized revenue) ---
    deal_agg_dict = {
        "hours": ("hours", "sum"),
        "billable_hours": ("billable_hours", "sum"),
        "staff_cost": ("work_cost", "sum"),
    }
    if "entry_revenue" in te.columns:
        deal_agg_dict["entry_revenue"] = ("entry_revenue", "sum")
    te_with_deal = te[te["deal_name"].notna() & (te["deal_name"] != "")]
    deals_summary = te_with_deal.groupby(
        ["deal_name", "service_type"]
    ).agg(**deal_agg_dict).reset_index()
    deals_summary["service_type"] = deals_summary["service_type"].fillna("(no type)")
    if "entry_revenue" not in deals_summary.columns:
        deals_summary["entry_revenue"] = 0

    # Use per-budget recognized revenue from financial_item_reports_by_budget
    fir_budget = load_financial_item_reports_by_budget()
    budget_map = _build_budget_to_deal_map()

    if fir_budget is not None and budget_map:
        # Sum revenue per deal across all months
        fir_budget["deal_name"] = fir_budget["budget_id"].map(
            lambda bid: budget_map.get(bid, {}).get("deal_name")
        )
        fir_budget_valid = fir_budget[fir_budget["deal_name"].notna()]
        deal_revenue = fir_budget_valid.groupby("deal_name").agg(
            budget_revenue=("revenue", "sum"),
            budget_invoiced=("invoiced", "sum"),
        ).reset_index()

        # Merge deal-level revenue, then distribute proportionally by billable_hours
        # across service_type rows within each deal (avoid duplicating total to each row)
        deals_summary = deals_summary.merge(deal_revenue, on="deal_name", how="left")
        deals_summary["budget_revenue"] = deals_summary["budget_revenue"].fillna(0)
        deals_summary["budget_invoiced"] = deals_summary["budget_invoiced"].fillna(0)

        # Proportional allocation: each service_type row gets share of deal revenue
        # based on its share of billable_hours within that deal
        deal_total_hours = deals_summary.groupby("deal_name")["billable_hours"].transform("sum")
        hour_share = (deals_summary["billable_hours"] / deal_total_hours.replace(0, float("nan"))).fillna(0)
        # For deals with 0 billable hours, distribute equally across rows
        rows_per_deal = deals_summary.groupby("deal_name")["deal_name"].transform("count")
        equal_share = 1.0 / rows_per_deal
        hour_share = hour_share.where(deal_total_hours > 0, equal_share)

        deals_summary["budget_revenue"] = (deals_summary["budget_revenue"] * hour_share).round(2)
        deals_summary["budget_invoiced"] = (deals_summary["budget_invoiced"] * hour_share).round(2)

        # Revenue = per-budget recognized revenue (covers both T&M and fixed-price)
        # Fall back to entry_revenue only for deals not found in budget data
        deals_summary["revenue"] = deals_summary["budget_revenue"].where(
            deals_summary["budget_revenue"] > 0,
            deals_summary["entry_revenue"],
        )
        print(f"  Enriched deals_summary with per-budget revenue ({len(deal_revenue)} deals matched)")
    else:
        deals_summary["revenue"] = deals_summary.get("entry_revenue", 0)
        deals_summary["budget_revenue"] = 0
        deals_summary["budget_invoiced"] = 0

    # --- Overbudget deals ---
    flaggable = client_budgets[
        (client_budgets["budgeted_hours"] > 0) | (client_budgets["budget_total"] > 0)
    ].copy()
    def flag_deal(row):
        hp = row["worked_hours"] / row["budgeted_hours"] * 100 if row["budgeted_hours"] > 0 else 0
        cp = row["cost"] / row["budget_total"] * 100 if row["budget_total"] > 0 else 0
        if hp > 100 or cp > 100:
            return "RED"
        elif hp > 70 or cp > 70:
            return "AMBER"
        return "GREEN"
    flaggable["flag"] = flaggable.apply(flag_deal, axis=1)
    flaggable["hours_burn_pct"] = (
        flaggable["worked_hours"] / flaggable["budgeted_hours"].replace(0, float("nan")) * 100
    ).round(1)
    flaggable["cost_burn_pct"] = (
        flaggable["cost"] / flaggable["budget_total"].replace(0, float("nan")) * 100
    ).round(1)
    flaggable["overspend_cost"] = (flaggable["cost"] - flaggable["budget_total"]).clip(lower=0).round(0)
    flaggable["overspend_hours"] = (flaggable["worked_hours"] - flaggable["budgeted_hours"]).clip(lower=0).round(1)
    ob_cols = ["name", "company_name", "company_id", "project_id", "id",
               "flag", "hours_burn_pct", "cost_burn_pct",
               "worked_hours", "budgeted_hours", "revenue", "cost",
               "budget_total", "profit", "overspend_cost", "overspend_hours"]
    ob_cols = [c for c in ob_cols if c in flaggable.columns]
    overbudget = flaggable[ob_cols]

    # --- Client summary ---
    deal_fin = client_budgets.groupby("company_name").agg(
        revenue=("revenue", "sum"),
        deal_count=("id", "count"),
    ).reset_index().rename(columns={"company_name": "client_name"})
    # Get company_id and project_id per client (first non-null from deals)
    _cli_ids = client_budgets.dropna(subset=["company_id"]).groupby("company_name").agg(
        company_id=("company_id", "first"),
        project_id=("project_id", "first"),
    ).reset_index().rename(columns={"company_name": "client_name"})
    deal_fin = deal_fin.merge(_cli_ids, on="client_name", how="left")
    te_client = te[te["client_name"].notna()].groupby("client_name").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("work_cost", "sum"),
        people=("person_name", "nunique"),
    ).reset_index()
    clients = te_client.merge(deal_fin, on="client_name", how="left")
    clients["deal_revenue"] = clients["revenue"].fillna(0)  # keep deal-level as fallback
    clients["revenue"] = clients["deal_revenue"]  # will be overridden with recognized rev later
    # Contingent columns — remaining values from Excel (source of truth)
    clients["has_contingent"] = clients["client_name"].isin(_contingent_company_set)
    clients["contingent_remaining"] = clients["client_name"].map(
        lambda n: _contingent_remaining.get(n, 0)
    )

    # Overbudget count per client
    ob_by_client = {}
    for _, row in flaggable.iterrows():
        cname = row.get("company_name", "")
        if cname not in ob_by_client:
            ob_by_client[cname] = 0
        hp = row["worked_hours"] / row["budgeted_hours"] * 100 if row["budgeted_hours"] > 0 else 0
        cp = row["cost"] / row["budget_total"] * 100 if row["budget_total"] > 0 else 0
        if hp > 70 or cp > 70:
            ob_by_client[cname] += 1
    clients["overbudget_deals"] = clients["client_name"].map(ob_by_client).fillna(0).astype(int)
    clients = clients.sort_values("revenue", ascending=False)

    # --- Hygiene monthly (exclude 0h entries from note compliance counts) ---
    te_nonzero = te[te["hours"] > 0]
    hygiene_monthly = te_nonzero.groupby("month").agg(
        total=("id", "count"),
        with_note=("has_note", "sum"),
    ).reset_index()
    hygiene_monthly["month"] = hygiene_monthly["month"].astype(str)
    hygiene_monthly["pct_with_note"] = (
        hygiene_monthly["with_note"] / hygiene_monthly["total"] * 100
    ).round(1)

    # --- Hygiene per person ---
    hygiene_person = te_nonzero.groupby(["person_name", "team"]).agg(
        total=("id", "count"),
        missing_notes=("has_note", lambda x: (~x).sum()),
    ).reset_index()
    hygiene_person["note_pct"] = (
        (hygiene_person["total"] - hygiene_person["missing_notes"])
        / hygiene_person["total"] * 100
    ).round(1)

    # --- Person scorecard (person × month) ---
    # Note compliance from non-zero entries only; hours/billable from all entries
    person_monthly_hyg_notes = te_nonzero.groupby(["person_name", "team", "month"]).agg(
        total=("id", "count"),
        missing_notes=("has_note", lambda x: (~x).sum()),
    ).reset_index()
    person_monthly_hyg_hours = te.groupby(["person_name", "team", "month"]).agg(
        hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
    ).reset_index()
    person_monthly_hyg = person_monthly_hyg_notes.merge(
        person_monthly_hyg_hours, on=["person_name", "team", "month"], how="outer"
    ).fillna(0)
    person_monthly_hyg["month"] = person_monthly_hyg["month"].astype(str)
    person_monthly_hyg["note_pct"] = (
        (person_monthly_hyg["total"] - person_monthly_hyg["missing_notes"])
        / person_monthly_hyg["total"].replace(0, 1) * 100
    ).round(1)
    person_monthly_hyg["util_pct"] = (
        person_monthly_hyg["billable_hours"] / person_monthly_hyg["hours"] * 100
    ).fillna(0).round(1)

    # --- Billability: individual non-billable entries ---
    nb = te[(te["is_billable"] == False) & (te["hours"] > 0)].copy()
    if not nb.empty:
        billability_entries = nb[["person_name", "team", "date", "hours",
                                  "deal_name", "client_name", "service_type", "note"]].copy()
        billability_entries["date"] = billability_entries["date"].dt.strftime("%Y-%m-%d")
        billability_entries["hours"] = billability_entries["hours"].round(2)
        billability_entries["deal_name"] = billability_entries["deal_name"].fillna("(no deal)")
        billability_entries["client_name"] = billability_entries["client_name"].fillna("(no client)")
        billability_entries["service_type"] = billability_entries["service_type"].fillna("(no type)")
        billability_entries["note"] = billability_entries["note"].fillna("")
    else:
        billability_entries = pd.DataFrame()

    # --- Budget audit (deals with missing/placeholder budgets) ---
    budget_audit = []
    for _, row in client_budgets.iterrows():
        bh = row.get("budgeted_hours", 0) or 0
        bt = row.get("budget_total", 0) or 0
        wh = row.get("worked_hours", 0) or 0
        name = row.get("name", "")
        company = row.get("company_name", "")

        # T&M retainers ("Ongoing"/"Contingent") are intentionally unbudgeted — skip
        name_lower = name.lower()
        is_retainer = "ongoing" in name_lower or "contingent" in name_lower

        if bh == 0 and bt == 0 and not is_retainer:
            issue = "no_budget"
        elif bh > 0 and bh <= 2 and wh > 10:
            issue = "placeholder"
        else:
            continue

        budget_audit.append({
            "name": name,
            "company_name": company,
            "deal_id": row.get("id", None),
            "company_id": row.get("company_id", None),
            "project_id": row.get("project_id", None),
            "issue": issue,
            "worked_hours": round(wh, 1),
            "budgeted_hours": round(bh, 1),
            "budget_total": round(bt, 0),
            "revenue": round(row.get("revenue", 0) or 0, 0),
            "cost": round(row.get("cost", 0) or 0, 0),
        })
    budget_audit.sort(key=lambda x: -x["worked_hours"])

    # --- Missing service type ---
    missing_stype = []
    no_stype = te[te["service_type"].isna()]
    if not no_stype.empty:
        for deal_name, grp in no_stype.groupby("deal_name"):
            missing_stype.append({
                "deal_name": deal_name or "(no deal)",
                "entries": len(grp),
                "hours": round(grp["hours"].sum(), 1),
            })
        missing_stype.sort(key=lambda x: -x["hours"])

    # --- Load services for budget health checks ---
    services_df = transform_endpoint("services")

    # --- Build activity date range per deal (for timeframe filtering) ---
    deal_activity = {}
    if "deal_id" in te.columns and "month" in te.columns:
        for did, grp in te.groupby("deal_id"):
            months = grp["month"].dropna().astype(str)
            if not months.empty:
                deal_activity[did] = {
                    "first_month": months.min()[:7],
                    "last_month": months.max()[:7],
                }

    # --- PSO Health Checks ---
    pso_health = []
    if services_df is not None:
        # Group services by deal
        svcs_by_deal = {}
        for _, svc in services_df.iterrows():
            did = svc.get("deal_id")
            if not did:
                continue
            svcs_by_deal.setdefault(did, []).append(svc)

        for _, deal in client_budgets.iterrows():
            name = str(deal.get("name", ""))
            if "PSO" not in name.upper() and "pso" not in name.lower():
                continue
            did = deal.get("id", "")
            deal_svcs = svcs_by_deal.get(did, [])
            company = deal.get("company_name", "")
            worked = deal.get("worked_hours", 0) or 0
            budgeted = deal.get("budgeted_hours", 0) or 0
            revenue = deal.get("revenue", 0) or 0
            cost = deal.get("cost", 0) or 0
            if budgeted > 0:
                burn_pct = round(worked / budgeted * 100, 1)
            elif worked > 0:
                burn_pct = 999.9
            else:
                burn_pct = 0

            issues = []
            if budgeted == 0 and worked > 0:
                issues.append("no_budget_set")
            # Check: missing overspend service
            has_overspend = any(
                "overspend" in str(s.get("name", "")).lower() or "overspent" in str(s.get("name", "")).lower()
                for s in deal_svcs
            )
            if not has_overspend:
                issues.append("missing_overspend")

            # Check: wrong service count (should be exactly 2)
            active_svcs = [s for s in deal_svcs if not s.get("deleted_at")]
            if len(active_svcs) != 2:
                issues.append(f"wrong_svc_count ({len(active_svcs)})")

            # Check: overspend marked billable
            for s in deal_svcs:
                sname = str(s.get("name", "")).lower()
                if ("overspend" in sname or "overspent" in sname) and s.get("billing_type_id") != 3:
                    issues.append("overspend_billable")
                    break

            # Check: overbudget
            if burn_pct > 100:
                issues.append("overbudget")

            # Check: budget cap exceeded (revenue > budget_total)
            for s in deal_svcs:
                if s.get("budget_cap_enabled"):
                    svc_rev = float(s.get("revenue", 0) or 0) / 100.0
                    svc_qty = float(s.get("quantity", 0) or 0)
                    svc_price = float(s.get("price", 0) or 0) / 100.0
                    cap_value = svc_qty * svc_price
                    if cap_value > 0 and svc_rev >= cap_value * 0.9:
                        cap_pct = round(svc_rev / cap_value * 100, 1)
                        issues.append(f"cap_hit ({cap_pct}%)")
                        break

            act = deal_activity.get(did, {})
            pso_health.append({
                "deal_id": str(did),
                "name": name,
                "company_name": company,
                "worked_hours": round(worked, 1),
                "budgeted_hours": round(budgeted, 1),
                "burn_pct": burn_pct,
                "revenue": round(revenue, 0),
                "cost": round(cost, 0),
                "service_count": len(active_svcs),
                "has_overspend": has_overspend,
                "issues": ", ".join(issues) if issues else "",
                "issue_count": len(issues),
                "severity": "critical" if "overspend_billable" in str(issues) else
                            "warning" if issues else "ok",
                "closed_at": str(deal.get("closed_at", ""))[:10] if deal.get("closed_at") and str(deal.get("closed_at")) != "nan" else "",
                "first_month": act.get("first_month", ""),
                "last_month": act.get("last_month", ""),
            })
        pso_health.sort(key=lambda x: (-x["issue_count"], -x["burn_pct"]))

    # --- Missing Overspend Service (non-PSO, non-retainer budgets) ---
    missing_overspend = []
    if services_df is not None:
        for _, deal in client_budgets.iterrows():
            did = deal.get("id", "")
            name = str(deal.get("name", ""))
            name_lower = name.lower()
            # Skip PSO (covered above), internal, and ongoing/contingent retainers
            if "PSO" in name.upper():
                continue
            if "ongoing" in name_lower or "contingent" in name_lower:
                continue
            company = deal.get("company_name", "")
            if company and company.lower() == "leadstreet":
                continue

            deal_svcs = svcs_by_deal.get(did, [])
            # Only flag budgets that have billable T&M services
            has_tm_billable = any(
                s.get("billing_type_id") == 2 and s.get("billable")
                for s in deal_svcs
            )
            if not has_tm_billable:
                continue

            has_overspend = any(
                ("overspend" in str(s.get("name", "")).lower() or "overspent" in str(s.get("name", "")).lower())
                and s.get("billing_type_id") == 3
                for s in deal_svcs
            )
            if has_overspend:
                continue

            # Check for billable overspend (wrong config)
            has_billable_overspend = any(
                ("overspend" in str(s.get("name", "")).lower() or "overspent" in str(s.get("name", "")).lower())
                and s.get("billing_type_id") != 3
                for s in deal_svcs
            )

            worked = deal.get("worked_hours", 0) or 0
            svc_count = len([s for s in deal_svcs if not s.get("deleted_at")])

            missing_overspend.append({
                "name": name,
                "company_name": company,
                "deal_id": str(did),
                "project_id": str(deal.get("project_id", "")),
                "worked_hours": round(worked, 1),
                "revenue": round(deal.get("revenue", 0) or 0, 0),
                "service_count": svc_count,
                "issue": "overspend_billable" if has_billable_overspend else "missing",
            })
        missing_overspend.sort(key=lambda x: -x["service_count"])

    # --- Closed budgets with post-close activity ---
    closed_with_activity = []
    te_by_deal = te_with_deal.groupby("deal_name")
    for _, deal in client_budgets.iterrows():
        closed_at = deal.get("closed_at")
        if not closed_at:
            continue
        name = str(deal.get("name", ""))
        company = deal.get("company_name", "")
        try:
            closed_date = pd.to_datetime(closed_at).tz_localize(None)
        except Exception:
            closed_date = pd.to_datetime(closed_at, utc=True).tz_localize(None)

        if name in te_by_deal.groups:
            grp = te_by_deal.get_group(name)
            if "date" in grp.columns:
                post_close = grp[grp["date"] > closed_date]
                if not post_close.empty:
                    closed_with_activity.append({
                        "name": name,
                        "company_name": company,
                        "deal_id": str(deal.get("id", "")),
                        "project_id": str(deal.get("project_id", "")),
                        "closed_date": str(closed_date.date()),
                        "post_close_hours": round(post_close["hours"].sum(), 1),
                        "post_close_entries": len(post_close),
                        "post_close_revenue": round(
                            post_close["entry_revenue"].sum() if "entry_revenue" in post_close.columns else 0, 0
                        ),
                        "last_entry": str(post_close["date"].max().date()) if not post_close.empty else "",
                    })
    closed_with_activity.sort(key=lambda x: -x["post_close_hours"])

    # --- Stale budgets (open, no recent activity, should be delivered) ---
    stale_budgets = []
    for _, deal in client_budgets.iterrows():
        # Only open budgets (not closed, not delivered)
        if deal.get("closed_at") or deal.get("delivered_on"):
            continue
        name = str(deal.get("name", ""))
        company = deal.get("company_name", "")
        if company and company.lower() == "leadstreet":
            continue
        # Skip ongoing/contingent retainers (expected to stay open)
        name_lower = name.lower()
        if "ongoing" in name_lower or "contingent" in name_lower:
            continue

        worked = deal.get("worked_hours", 0) or 0
        budgeted = deal.get("budgeted_hours", 0) or 0

        # Check days since last activity
        days_inactive = deal.get("days_since_last_activity")
        if days_inactive is None or pd.isna(days_inactive):
            continue
        days_inactive = int(days_inactive)

        # Flag if: >30 days inactive AND has had some work done
        if days_inactive > 30 and worked > 0:
            burn_pct = round(worked / budgeted * 100, 1) if budgeted > 0 else 0
            last_act = deal.get("last_activity_at")
            last_act_date = str(last_act)[:10] if last_act and str(last_act) != "nan" else ""
            stale_budgets.append({
                "name": name,
                "company_name": company,
                "deal_id": str(deal.get("id", "")),
                "project_id": str(deal.get("project_id", "")),
                "worked_hours": round(worked, 1),
                "budgeted_hours": round(budgeted, 1),
                "burn_pct": burn_pct,
                "days_inactive": days_inactive,
                "last_activity_at": last_act_date,
                "revenue": round(deal.get("revenue", 0) or 0, 0),
                "projected_revenue": round(deal.get("projected_revenue", 0) or 0, 0),
                "severity": "critical" if (burn_pct > 90 and days_inactive > 60) else
                            "warning" if (burn_pct > 70 or days_inactive > 90) else "info",
            })
    stale_budgets.sort(key=lambda x: (-x["burn_pct"], -x["days_inactive"]))

    # --- Data health summary counts ---
    total_entries = len(te)
    data_health_summary = {
        "total_entries": total_entries,
        "missing_notes": int((~te["has_note"]).sum()),
        "missing_notes_pct": round((~te["has_note"]).sum() / total_entries * 100, 1),
        "zero_hour_entries": int((te["hours"] == 0).sum()),
        "missing_service_type": int(te["service_type"].isna().sum()),
        "budget_no_limit": len([a for a in budget_audit if a["issue"] == "no_budget"]),
        "budget_placeholder": len([a for a in budget_audit if a["issue"] == "placeholder"]),
        "missing_overspend_count": len(missing_overspend),
        "overbudget_red_real": len([
            r for r in overbudget.to_dict("records")
            if r.get("flag") == "RED"
            and (r.get("budgeted_hours", 0) or 0) > 2
        ]),
        "overbudget_red_total": len([
            r for r in overbudget.to_dict("records")
            if r.get("flag") == "RED"
        ]),
        "pso_total": len(pso_health),
        "pso_with_issues": len([p for p in pso_health if p["issue_count"] > 0]),
        "pso_overbudget": len([p for p in pso_health if p["burn_pct"] > 100]),
        "missing_overspend_total": len(missing_overspend),
        "closed_with_activity": len(closed_with_activity),
        "closed_activity_hours": round(sum(c["post_close_hours"] for c in closed_with_activity), 1),
        "stale_budgets_total": len(stale_budgets),
        "stale_critical": len([s for s in stale_budgets if s["severity"] == "critical"]),
    }

    # --- Budget data ---
    budgets = load_all_budgets()

    # Enrich financial_monthly with budget targets
    budget_monthly = []
    for year, b in budgets.items():
        if b:
            for m in b["monthly"]:
                budget_monthly.append({
                    "month": m["month"],
                    "budget_revenue": m["revenue_hours"],
                    "budget_people_cost": m.get("people_cost", 0),
                    "budget_margin": m.get("gross_margin_people", m["gross_margin"]),
                    "budget_margin_pct": m.get("margin_pct_people", m["margin_pct"]),
                })

    # Enrich people with billability + rate targets
    productive_names = list(set(r for r in te["person_name"].unique() if r))
    budget_people = []
    for year, b in budgets.items():
        if b:
            for p in b["people"]:
                matched = match_budget_name(p["name"], productive_names)
                budget_people.append({
                    "budget_name": p["name"],
                    "productive_name": matched,
                    "billability_target": p["billability_target"],
                    "rate_target": p["rate_target"],
                    "is_new_hire": p.get("is_new_hire", False),
                    "year": year,
                })

    # --- Budget scenarios for Executive Summary ---
    budget_current = []
    budget_with_new_hires = []
    b2026 = budgets.get(2026)
    if b2026:
        for m in b2026["monthly"]:
            budget_current.append({
                "month": m["month"],
                "revenue": m["revenue_hours"],
                "people_cost": m["people_cost"],
                "gross_margin": m["gross_margin_people"],
                "margin_pct": m["margin_pct_people"],
            })
        for m in b2026.get("monthly_with_new_hires", []):
            budget_with_new_hires.append({
                "month": m["month"],
                "revenue": m["revenue_hours"],
                "people_cost": m["people_cost"],
                "gross_margin": m["gross_margin_people"],
                "margin_pct": m["margin_pct_people"],
            })

    # --- Service type monthly (service_type × month) ---
    # Start with hours/cost from time entries
    stype_monthly_agg = {
        "hours": ("hours", "sum"),
        "billable_hours": ("billable_hours", "sum"),
        "staff_cost": ("work_cost", "sum"),
    }
    if "entry_revenue" in te.columns:
        stype_monthly_agg["entry_revenue"] = ("entry_revenue", "sum")
    stype_monthly = te.groupby(["service_type", "month"]).agg(
        **stype_monthly_agg
    ).reset_index()
    stype_monthly["month"] = stype_monthly["month"].astype(str)
    stype_monthly["service_type"] = stype_monthly["service_type"].fillna("(no type)")
    if "entry_revenue" not in stype_monthly.columns:
        stype_monthly["entry_revenue"] = 0

    # Enrich with recognized revenue by service type
    # Prefer direct service_type grouping from API (accurate), fall back to budget-based (legacy)
    fir_by_stype = load_financial_item_reports_by_service_type()
    if fir_by_stype is not None:
        stype_monthly = stype_monthly.merge(
            fir_by_stype.rename(columns={"month_str": "month", "revenue": "api_revenue"}),
            on=["service_type", "month"], how="left"
        )
        stype_monthly["api_revenue"] = stype_monthly["api_revenue"].fillna(0)
        stype_monthly["revenue"] = stype_monthly["api_revenue"].where(
            stype_monthly["api_revenue"] > 0,
            stype_monthly["entry_revenue"],
        )
    elif fir_budget is not None and budget_map:
        # Legacy fallback: map budget → single service_type (less accurate for mixed deals)
        fir_budget["service_type"] = fir_budget["budget_id"].map(
            lambda bid: budget_map.get(bid, {}).get("service_type")
        )
        fir_budget["service_type"] = fir_budget["service_type"].fillna("(no type)")
        stype_budget_rev = fir_budget.groupby(["service_type", "month_str"]).agg(
            budget_revenue=("revenue", "sum"),
        ).reset_index().rename(columns={"month_str": "month"})
        stype_monthly = stype_monthly.merge(
            stype_budget_rev, on=["service_type", "month"], how="left"
        )
        stype_monthly["budget_revenue"] = stype_monthly["budget_revenue"].fillna(0)
        stype_monthly["revenue"] = stype_monthly["budget_revenue"].where(
            stype_monthly["budget_revenue"] > 0,
            stype_monthly["entry_revenue"],
        )
    else:
        stype_monthly["revenue"] = stype_monthly["entry_revenue"]

    # --- Client monthly (client_name × month) ---
    client_monthly_agg = {
        "hours": ("hours", "sum"),
        "billable_hours": ("billable_hours", "sum"),
        "staff_cost": ("work_cost", "sum"),
    }
    if "entry_revenue" in te.columns:
        client_monthly_agg["entry_revenue"] = ("entry_revenue", "sum")
    te_with_client = te[te["client_name"].notna()]
    client_monthly = te_with_client.groupby(["client_name", "month"]).agg(
        **client_monthly_agg
    ).reset_index()
    client_monthly["month"] = client_monthly["month"].astype(str)
    if "entry_revenue" not in client_monthly.columns:
        client_monthly["entry_revenue"] = 0

    # Enrich client_monthly with company_id and project_id for deep links
    if "company_id" in clients.columns:
        cli_id_map = clients[["client_name", "company_id", "project_id"]].drop_duplicates("client_name")
        client_monthly = client_monthly.merge(
            cli_id_map, on="client_name", how="left"
        )

    # Enrich with per-budget recognized revenue by client
    if fir_budget is not None and budget_map:
        fir_budget["company_name"] = fir_budget["budget_id"].map(
            lambda bid: budget_map.get(bid, {}).get("company_name")
        )
        fir_budget_valid_cli = fir_budget[fir_budget["company_name"].notna()]
        cli_budget_rev = fir_budget_valid_cli.groupby(["company_name", "month_str"]).agg(
            budget_revenue=("revenue", "sum"),
        ).reset_index().rename(columns={"company_name": "client_name", "month_str": "month"})
        client_monthly = client_monthly.merge(
            cli_budget_rev, on=["client_name", "month"], how="left"
        )
        client_monthly["budget_revenue"] = client_monthly["budget_revenue"].fillna(0)
        client_monthly["revenue"] = client_monthly["budget_revenue"].where(
            client_monthly["budget_revenue"] > 0,
            client_monthly["entry_revenue"],
        )
        # Also override clients snapshot with recognized revenue
        cli_total_rev = cli_budget_rev.groupby("client_name")["budget_revenue"].sum().reset_index()
        cli_total_rev.columns = ["client_name", "recognized_revenue"]
        clients = clients.merge(cli_total_rev, on="client_name", how="left")
        clients["recognized_revenue"] = clients["recognized_revenue"].fillna(0)
        clients["revenue"] = clients["recognized_revenue"].where(
            clients["recognized_revenue"] > 0, clients["deal_revenue"]
        )
    else:
        client_monthly["revenue"] = client_monthly["entry_revenue"]

    # --- Deal monthly (for date-filtering scope creep radar) ---
    te_with_deal_name = te[te["deal_name"].notna() & (te["deal_name"] != "")]
    deal_monthly = te_with_deal_name.groupby(["deal_name", "month"]).agg(
        hours=("hours", "sum"),
    ).reset_index()
    deal_monthly["month"] = deal_monthly["month"].astype(str)

    # Recompute client derived fields with final revenue
    clients["gross_margin"] = clients["revenue"] - clients["staff_cost"]
    clients["margin_pct"] = (
        clients["gross_margin"] / clients["revenue"].replace(0, float("nan")) * 100
    ).fillna(0).round(1)
    clients["acph"] = (clients["staff_cost"] / clients["total_hours"]).round(2)
    clients["arph"] = (
        clients["revenue"] / clients["billable_hours"].replace(0, float("nan"))
    ).fillna(0).round(2)
    # Recompute contingent adjusted fields with final revenue
    # contingent_remaining = unearned balance from Excel (source of truth)
    # adjusted_revenue = revenue minus unearned contingent remaining
    clients["contingent_excess"] = clients["contingent_remaining"].clip(lower=0)
    clients["adjusted_revenue"] = (clients["revenue"] - clients["contingent_excess"]).clip(lower=0)
    clients["adjusted_gross_margin"] = clients["adjusted_revenue"] - clients["staff_cost"]
    clients["adjusted_margin"] = (
        clients["adjusted_gross_margin"]
        / clients["adjusted_revenue"].replace(0, float("nan")) * 100
    ).fillna(0).round(1)
    # Cap adjusted margin: when adjusted_revenue is tiny, margin % is meaningless
    clients.loc[
        clients["has_contingent"] & (clients["adjusted_gross_margin"] < 0),
        "adjusted_margin"
    ] = 0.0

    # --- Enrich people_monthly with recognized revenue ---
    # Allocate per-deal-per-month revenue to people proportionally by billable hours
    if fir_budget is not None and budget_map:
        # Build deal_name → month → recognized revenue
        fir_budget["deal_name"] = fir_budget["budget_id"].map(
            lambda bid: budget_map.get(bid, {}).get("deal_name")
        )
        deal_month_rev = fir_budget[fir_budget["deal_name"].notna()].groupby(
            ["deal_name", "month_str"]
        )["revenue"].sum().reset_index()
        deal_month_rev.columns = ["deal_name", "month", "deal_revenue"]

        # Get each person's share of billable hours per deal per month
        te_with_deal = te[te["deal_name"].notna() & (te["deal_name"] != "")]
        person_deal_month = te_with_deal.groupby(
            ["person_name", "deal_name", "month"]
        ).agg(
            person_bill=("billable_hours", "sum"),
        ).reset_index()
        person_deal_month["month"] = person_deal_month["month"].astype(str)

        # Total billable hours per deal per month
        deal_month_total = person_deal_month.groupby(
            ["deal_name", "month"]
        )["person_bill"].transform("sum").replace(0, float("nan"))
        person_deal_month["share"] = (person_deal_month["person_bill"] / deal_month_total).fillna(0)

        # Merge with deal revenue and compute person's share
        person_deal_month = person_deal_month.merge(
            deal_month_rev, on=["deal_name", "month"], how="left"
        )
        person_deal_month["deal_revenue"] = person_deal_month["deal_revenue"].fillna(0)
        person_deal_month["person_revenue"] = person_deal_month["share"] * person_deal_month["deal_revenue"]

        # Sum per person per month
        person_rev = person_deal_month.groupby(
            ["person_name", "month"]
        )["person_revenue"].sum().reset_index()
        person_rev.columns = ["person_name", "month", "recognized_revenue"]

        people_monthly = people_monthly.merge(
            person_rev, on=["person_name", "month"], how="left"
        )
        people_monthly["recognized_revenue"] = people_monthly["recognized_revenue"].fillna(0)
        people_monthly["revenue"] = people_monthly["recognized_revenue"].where(
            people_monthly["recognized_revenue"] > 0,
            people_monthly["entry_revenue"],
        )
    else:
        people_monthly["revenue"] = people_monthly["entry_revenue"]

    # --- Reconcile people_monthly revenue with financial_monthly ---
    # Some revenue can't be allocated to individuals (unmapped deals/budgets).
    # Distribute the gap proportionally by hours so totals match.
    fin_rev_by_month = financial_monthly.set_index("month")["revenue"].to_dict()
    for m, fin_rev in fin_rev_by_month.items():
        mask = people_monthly["month"] == m
        ppl_rev = people_monthly.loc[mask, "revenue"].sum()
        gap = fin_rev - ppl_rev
        if abs(gap) > 1 and people_monthly.loc[mask, "hours"].sum() > 0:
            month_hours = people_monthly.loc[mask, "hours"]
            total_hours = month_hours.sum()
            people_monthly.loc[mask, "revenue"] += gap * (month_hours / total_hours)

    # --- Variance drivers (2026 only, using recognized revenue) ---
    target_rate = b2026["target_rate"] if b2026 else 112.0

    # a) Service type variance — uses enriched stype_monthly (recognized revenue)
    stype_2026_monthly = stype_monthly[stype_monthly["month"] >= "2026-01"]
    stype_2026 = stype_2026_monthly.groupby("service_type").agg(
        total_hours=("hours", "sum"),
        billable_hours=("billable_hours", "sum"),
        staff_cost=("staff_cost", "sum"),
        allocated_revenue=("revenue", "sum"),
    ).reset_index()
    stype_2026["arph"] = (
        stype_2026["allocated_revenue"] / stype_2026["billable_hours"].replace(0, float("nan"))
    ).fillna(0).round(2)
    stype_2026["acph"] = (stype_2026["staff_cost"] / stype_2026["total_hours"]).round(2)

    variance_service_mix = []
    for _, row in stype_2026.iterrows():
        arph = row.get("arph", 0)
        if pd.isna(arph):
            arph = 0
        acph = row.get("acph", 0)
        if pd.isna(acph):
            acph = 0
        bill_hrs = row.get("billable_hours", 0) or 0
        if bill_hrs < 10:
            continue
        variance_service_mix.append({
            "service_type": row["service_type"] or "(no type)",
            "arph": round(arph, 2),
            "acph": round(acph, 2),
            "target_arph": target_rate,
            "billable_hours": round(bill_hrs, 1),
            "revenue_impact": round((arph - target_rate) * bill_hrs, 0),
        })
    variance_service_mix.sort(key=lambda x: x["arph"], reverse=True)

    # b) Person variance — uses enriched people_monthly (recognized revenue)
    variance_utilisation = []
    if b2026:
        budget_targets = {}
        for p in b2026["people"]:
            matched = match_budget_name(p["name"], productive_names)
            if matched:
                budget_targets[matched] = {
                    "billability_target": p["billability_target"],
                    "rate_target": p["rate_target"] or 112,
                }

        # Aggregate people_monthly for 2026 (already has recognized revenue)
        pm_2026 = people_monthly[people_monthly["month"] >= "2026-01"]
        person_agg_2026 = pm_2026.groupby("person_name").agg(
            total_hours=("hours", "sum"),
            billable_hours=("billable_hours", "sum"),
            staff_cost=("staff_cost", "sum"),
            allocated_revenue=("revenue", "sum"),
        ).reset_index()

        for _, row in person_agg_2026.iterrows():
            name = row["person_name"]
            if name not in budget_targets:
                continue
            tgt = budget_targets[name]
            actual_util = row["billable_hours"] / row["total_hours"] if row["total_hours"] > 0 else 0
            target_util = tgt["billability_target"] or 0
            actual_arph = (
                row["allocated_revenue"] / row["billable_hours"]
                if row["billable_hours"] > 0 else 0
            )
            target_arph = tgt["rate_target"]
            variance_utilisation.append({
                "person": name,
                "actual_util": round(actual_util * 100, 1),
                "target_util": round(target_util * 100, 1),
                "gap_pct": round((actual_util - target_util) * 100, 1),
                "actual_arph": round(actual_arph, 2),
                "target_arph": round(target_arph, 2),
                "total_hours": round(row["total_hours"], 1),
                "billable_hours": round(row["billable_hours"], 1),
                "revenue_impact": round((actual_util - target_util) * row["total_hours"] * target_arph, 0),
            })
        variance_utilisation.sort(key=lambda x: x["actual_util"])

    # Convert all to JSON-safe dicts
    def to_records(df):
        return json.loads(df.to_json(orient="records", date_format="iso"))

    return {
        "financial_monthly": to_records(financial_monthly),
        "people_monthly": to_records(people_monthly),
        "team_summary": to_records(team_summary),
        "service_type_monthly": to_records(stype_monthly),
        "deals_summary": to_records(deals_summary),
        "overbudget": to_records(overbudget),
        "clients": to_records(clients),
        "client_monthly": to_records(client_monthly),
        "hygiene_monthly": to_records(hygiene_monthly),
        "hygiene_person": to_records(hygiene_person),
        "person_scorecard": to_records(person_monthly_hyg),
        "billability_entries": to_records(billability_entries) if not billability_entries.empty else [],
        "budget_audit": budget_audit,
        "missing_stype": missing_stype,
        "data_health_summary": data_health_summary,
        "budget_monthly": budget_monthly,
        "budget_people": budget_people,
        "budget_target_rate": {
            year: (b["target_rate"] if b else None)
            for year, b in budgets.items()
        },
        "budget_current": budget_current,
        "budget_with_new_hires": budget_with_new_hires,
        "variance_service_mix": variance_service_mix,
        "variance_utilisation": variance_utilisation,
        "rate_card_actions": _compute_rate_card_actions(services_df),
        "pso_health": pso_health,
        "missing_overspend": missing_overspend,
        "closed_with_activity": closed_with_activity,
        "stale_budgets": stale_budgets,
        "deal_monthly": to_records(deal_monthly),
    }


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Run LeadStreet BI reports")
    parser.add_argument(
        "--report",
        choices=["financial", "people", "project", "client", "hygiene", "all"],
        default="all",
        help="Which report to run",
    )
    args = parser.parse_args()

    print("Loading data...")
    te = transform_endpoint("time_entries")
    deals = transform_endpoint("deals")

    if te is None:
        print("ERROR: No time_entries snapshot found. Run fetch.py first.")
        sys.exit(1)
    if deals is None:
        print("ERROR: No deals snapshot found. Run fetch.py first.")
        sys.exit(1)

    print(f"\nLoaded: {len(te):,} time entries, {len(deals):,} deals")

    reports = {
        "financial": lambda: report_financial(te, deals),
        "people": lambda: report_people(te),
        "project": lambda: report_project(deals, te),
        "client": lambda: report_client(te, deals),
        "hygiene": lambda: report_hygiene(te),
    }

    if args.report == "all":
        for name, func in reports.items():
            func()
    else:
        reports[args.report]()

    print(f"\nReports complete. CSVs + HTML charts saved to {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
