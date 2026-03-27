"""
load_budget.py — Parse LeadStreet budget Excel files for 2025 and 2026.

Each year has a different layout. Both parsers return a normalized structure:
{
    "monthly": [{"month": "2026-01", "revenue_hours": ..., "gross_margin": ..., "margin_pct": ...}, ...],
    "people": [{"name": "Andrew Llorera", "billability_target": 0.80, "rate_target": 112}, ...],
    "target_rate": 112.0  # blended target rate (None if not available)
}

Gracefully returns None if OneDrive is not synced or file is missing.
"""
from __future__ import annotations

from pathlib import Path

ONEDRIVE_BASE = Path.home() / (
    "Library/CloudStorage/OneDrive-Gedeeldebibliotheken-T-WESSC°BV/"
    "Storage-Meja - Documenten/Meja/Always On Group/Leadstreet/"
    "2. Finance/1. Budget&Forecast"
)


def _find_latest(directory: Path, pattern: str) -> Path | None:
    """Find the file with the highest date-prefix matching the glob pattern."""
    candidates = sorted(directory.glob(pattern), reverse=True)
    if candidates:
        print(f"  [auto-discover] Using: {candidates[0].name}")
    return candidates[0] if candidates else None


BUDGET_2026_FILE = _find_latest(ONEDRIVE_BASE / "2026", "*Leadstreet Budget 2026*.xlsx")
BUDGET_2025_FILE = _find_latest(ONEDRIVE_BASE / "2025", "*Leadstreet Budget 2025*.xlsx")


def load_budget_2026(path: Path | None = None) -> dict | None:
    """Parse 2026 budget Excel.

    BUD26 leadstreet sheet:
        Row 1: dates (cols B-M = Jan-Dec 2026)
        Row 3: TOTAL REVENUE FROM HOURS WORKED (cols B-M)
        Row 23: Gross margin (cols B-M)
        Row 24: Gross margin% (cols B-M, as decimal 0-1)

    Budget Timesheet-2026 sheet:
        Rows 9-26: current team members
        Col B: name, Col M: billability target, Col AB(28): rate target
        Rows 33-35: new hires
    """
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed, cannot load budget")
        return None

    path = path or BUDGET_2026_FILE
    if not path.exists():
        print(f"  Budget 2026 not found at {path}")
        return None

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        print(f"  ERROR loading budget 2026: {e}")
        return None

    # --- BUD26 leadstreet: monthly financials ---
    ws = wb["BUD26 leadstreet"]

    # Validate label
    label = str(ws.cell(row=3, column=1).value or "")
    if "TOTAL REVENUE FROM H" not in label:
        print(f"  WARNING: BUD26 row 3 label mismatch: '{label}'")

    monthly = []
    for col_idx in range(2, 14):  # B=2 through M=13 → Jan-Dec
        month_num = col_idx - 1  # col 2 → month 1
        month_str = f"2026-{month_num:02d}"

        rev = _num(ws.cell(row=3, column=col_idx).value)
        gm = _num(ws.cell(row=23, column=col_idx).value)
        gm_pct_raw = _num(ws.cell(row=24, column=col_idx).value)
        gm_pct = gm_pct_raw * 100 if gm_pct_raw and gm_pct_raw < 1 else gm_pct_raw

        # People cost = sum of rows 12-14 (contractors) + 16 (employees) + 17 (bestuurders)
        # This matches the dashboard's "Staff Cost" definition (salary-based, no overhead)
        people_cost = sum(
            _num(ws.cell(row=r, column=col_idx).value)
            for r in [12, 13, 14, 16, 17]
        )

        # Read "Gross margin only people" directly from spreadsheet (row 25-26)
        # These include all costs, not just the people_cost rows above
        gm_people = _num(ws.cell(row=25, column=col_idx).value)
        gm_pct_people_raw = _num(ws.cell(row=26, column=col_idx).value)
        gm_pct_people = gm_pct_people_raw * 100 if gm_pct_people_raw and gm_pct_people_raw < 1 else gm_pct_people_raw

        # Row 21: TOTAL COSTS
        total_costs = _num(ws.cell(row=21, column=col_idx).value)

        monthly.append({
            "month": month_str,
            "revenue_hours": rev,
            "people_cost": people_cost,
            "total_costs": total_costs,
            "gross_margin": gm,
            "gross_margin_people": gm_people,
            "margin_pct": round(gm_pct, 1) if gm_pct else 0,
            "margin_pct_people": round(gm_pct_people, 1) if gm_pct_people else 0,
        })

    # --- Budget Timesheet-2026: people targets ---
    ws2 = wb["Budget Timesheet-2026"]
    people_current = []
    for row in range(9, 27):  # rows 9-26
        name = ws2.cell(row=row, column=2).value
        if not name or str(name).strip() == "":
            continue
        bill = _num(ws2.cell(row=row, column=13).value)  # col M
        rate = _num(ws2.cell(row=row, column=28).value)  # col AB
        people_current.append({
            "name": str(name).strip(),
            "billability_target": bill,
            "rate_target": rate or 112,
            "is_new_hire": False,
        })

    people_new = []
    for row in range(33, 36):  # rows 33-35
        name = ws2.cell(row=row, column=2).value
        if not name or str(name).strip() == "":
            continue
        bill = _num(ws2.cell(row=row, column=13).value)
        rate = _num(ws2.cell(row=row, column=28).value)
        people_new.append({
            "name": str(name).strip(),
            "billability_target": bill,
            "rate_target": rate or 112,
            "is_new_hire": True,
        })

    # --- Budget Timesheet-2026: new hires monthly revenue (cols 30-41) + cost (cols 44-55) ---
    monthly_with_new_hires = []
    for month_idx in range(12):  # 0=Jan .. 11=Dec
        rev_col = 30 + month_idx   # cols 30-41 = revenue Jan-Dec
        cost_col = 44 + month_idx  # cols 44-55 = cost Jan-Dec
        new_hire_rev = 0.0
        new_hire_cost = 0.0
        for row in range(33, 36):  # rows 33-35: new hires
            new_hire_rev += _num(ws2.cell(row=row, column=rev_col).value)
            new_hire_cost += _num(ws2.cell(row=row, column=cost_col).value)

        base = monthly[month_idx]
        combined_rev = base["revenue_hours"] + new_hire_rev
        combined_people_cost = base["people_cost"] + new_hire_cost
        combined_gm = combined_rev - combined_people_cost

        monthly_with_new_hires.append({
            "month": base["month"],
            "revenue_hours": combined_rev,
            "people_cost": combined_people_cost,
            "gross_margin_people": combined_gm,
            "margin_pct_people": round(combined_gm / combined_rev * 100, 1) if combined_rev else 0,
        })

    wb.close()

    return {
        "year": 2026,
        "monthly": monthly,
        "monthly_with_new_hires": monthly_with_new_hires,
        "people": people_current,
        "people_new_hires": people_new,
        "target_rate": 112.0,
    }


def load_budget_2025(path: Path | None = None) -> dict | None:
    """Parse 2025 budget Excel.

    BUD25 sheet:
        Row 1: dates (cols N-Y = Jan-Dec 2025)
        Row 3: TOTAL REVENUE FROM HOURS WORKED (cols N-Y)
        Row 39: Gross margin (cols N-Y)
        Row 40: Gross margin% (cols N-Y, as decimal 0-1)

    Budget Timesheet-2025 sheet:
        Rows 9-28: team members
        Col B: name, Col L: billability target, Col AA(27): rate target
    """
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed, cannot load budget")
        return None

    path = path or BUDGET_2025_FILE
    if not path.exists():
        print(f"  Budget 2025 not found at {path}")
        return None

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        print(f"  ERROR loading budget 2025: {e}")
        return None

    # --- BUD25: monthly financials ---
    ws = wb["BUD25"]

    label = str(ws.cell(row=3, column=1).value or "")
    if "TOTAL REVENUE FROM HOURS" not in label:
        print(f"  WARNING: BUD25 row 3 label mismatch: '{label}'")

    monthly = []
    for col_idx in range(14, 26):  # N=14 through Y=25 → Jan-Dec
        month_num = col_idx - 13  # col 14 → month 1
        month_str = f"2025-{month_num:02d}"

        rev = _num(ws.cell(row=3, column=col_idx).value)
        gm = _num(ws.cell(row=39, column=col_idx).value)
        gm_pct_raw = _num(ws.cell(row=40, column=col_idx).value)
        gm_pct = gm_pct_raw * 100 if gm_pct_raw and gm_pct_raw < 1 else gm_pct_raw

        monthly.append({
            "month": month_str,
            "revenue_hours": rev,
            "gross_margin": gm,
            "margin_pct": round(gm_pct, 1) if gm_pct else 0,
        })

    # --- Budget Timesheet-2025: people targets ---
    ws2 = wb["Budget Timesheet-2025"]
    people = []
    for row in range(9, 29):  # rows 9-28
        name = ws2.cell(row=row, column=2).value
        if not name or str(name).strip() == "":
            continue
        # Skip placeholder entries like "Nieuwe PMO"
        name_str = str(name).strip()
        bill = _num(ws2.cell(row=row, column=12).value)  # col L
        rate = _num(ws2.cell(row=row, column=27).value)  # col AA
        people.append({
            "name": name_str,
            "billability_target": bill,
            "rate_target": rate,
            "is_new_hire": False,
        })

    wb.close()

    # 2025 has mixed rates (80 for most, 100 for owners) — no single blended target
    return {
        "year": 2025,
        "monthly": monthly,
        "people": people,
        "people_new_hires": [],
        "target_rate": None,
    }


def load_all_budgets() -> dict[int, dict | None]:
    """Load budgets for all years. Returns {year: budget_dict_or_None}."""
    result = {}
    for year, loader in [(2025, load_budget_2025), (2026, load_budget_2026)]:
        try:
            result[year] = loader()
            if result[year]:
                print(f"  Budget {year}: loaded {len(result[year]['monthly'])} months, "
                      f"{len(result[year]['people'])} people")
        except Exception as e:
            print(f"  ERROR loading budget {year}: {e}")
            result[year] = None
    return result


def _num(val) -> float:
    """Safely convert Excel cell value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# --- Name matching between budget and Productive ---
# Budget uses informal/short names; Productive uses full names from the system.
# This mapping resolves known mismatches.
BUDGET_TO_PRODUCTIVE_NAME = {
    "Chloe": "Chloe De Smet",
    "Nieuwe PMO": None,  # placeholder, skip
    "Project manager": None,  # new hire role, skip
    "Sales /CSM": None,  # new hire role, skip
    "Developer": None,  # new hire role, skip
}


def match_budget_name(budget_name: str, productive_names: list[str]) -> str | None:
    """Match a budget name to a Productive person name.

    Tries: exact match → mapped name → first-name match → substring match.
    Returns None if no match found.
    """
    # Check explicit mapping first
    if budget_name in BUDGET_TO_PRODUCTIVE_NAME:
        return BUDGET_TO_PRODUCTIVE_NAME[budget_name]

    # Exact match
    if budget_name in productive_names:
        return budget_name

    # First name match (budget often uses just first name or short form)
    budget_first = budget_name.split()[0].lower() if budget_name else ""
    for pname in productive_names:
        if pname.lower().startswith(budget_first):
            return pname

    # Substring match
    budget_lower = budget_name.lower()
    for pname in productive_names:
        if budget_lower in pname.lower() or pname.lower() in budget_lower:
            return pname

    return None


if __name__ == "__main__":
    import json
    budgets = load_all_budgets()
    for year, b in budgets.items():
        if b:
            print(f"\n{'='*60}")
            print(f"Budget {year}")
            print(f"{'='*60}")
            print(f"Target rate: {b['target_rate']}")
            print(f"\nMonthly revenue targets:")
            for m in b["monthly"]:
                print(f"  {m['month']}: rev={m['revenue_hours']:>10,.0f}  "
                      f"GM={m['gross_margin']:>10,.0f}  margin={m['margin_pct']:.1f}%")
            print(f"\nPeople targets:")
            for p in b["people"]:
                print(f"  {p['name']:<35} bill={p['billability_target']:.0%}  "
                      f"rate={p['rate_target']}")
            if b.get("people_new_hires"):
                print(f"\nNew hires:")
                for p in b["people_new_hires"]:
                    print(f"  {p['name']:<35} bill={p['billability_target']:.0%}  "
                          f"rate={p['rate_target']}")
